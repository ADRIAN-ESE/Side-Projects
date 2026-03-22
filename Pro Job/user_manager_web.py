"""
User Manager Pro — Flask Web App
=================================
Merged & upgraded from user_manager_gui.py + user_collector.py

Features:
  - Full SPA (single-page app) embedded in one file
  - Dashboard with live statistics
  - Add / Edit / Delete users
  - Search & filter
  - Promotion request & approval workflow
  - Import from JSON / CSV
  - Export to JSON / CSV / Excel
  - Dark industrial aesthetic with monospace UI

Requirements:
  pip install flask openpyxl

Run:
  python user_manager_web.py
  Then open: http://localhost:5000
"""

from flask import Flask, request, jsonify, render_template_string, send_file
import json
import os
import csv
import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── Config ────────────────────────────────────────────────────────────────────
DATA_FILE = "users.json"
app = Flask(__name__)


# ── Data helpers ──────────────────────────────────────────────────────────────
def load_users() -> dict:
    if not os.path.exists(DATA_FILE):
        return {}
    try:
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return {}


def save_users(users: dict) -> None:
    with open(DATA_FILE, "w") as f:
        json.dump(users, f, indent=4)


# ═════════════════════════════════════════════════════════════════════════════
#  HTML TEMPLATE
# ═════════════════════════════════════════════════════════════════════════════
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>User Manager Pro</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@300;400;500;700&family=Syne:wght@400;600;700;800&display=swap" rel="stylesheet">
<style>
/* ── Reset & Tokens ─────────────────────────────────────────────────────── */
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

:root {
  --bg:        #0d0f1a;
  --surface:   #131625;
  --card:      #1a1d30;
  --border:    #252842;
  --accent:    #4f8ef7;
  --accent2:   #7c5df9;
  --success:   #3dba7a;
  --warning:   #e09a30;
  --danger:    #e05252;
  --text:      #d4d8f0;
  --text-muted:#6e7494;
  --mono:      'JetBrains Mono', monospace;
  --sans:      'Syne', sans-serif;
  --radius:    10px;
  --sidebar-w: 220px;
}

html, body { height: 100%; background: var(--bg); color: var(--text); font-family: var(--mono); font-size: 13px; }

/* ── Scrollbar ──────────────────────────────────────────────────────────── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: var(--bg); }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #3a3f60; }

/* ── Layout ─────────────────────────────────────────────────────────────── */
#app { display: flex; height: 100vh; overflow: hidden; }

/* ── Sidebar ────────────────────────────────────────────────────────────── */
#sidebar {
  width: var(--sidebar-w); flex-shrink: 0;
  background: var(--surface);
  border-right: 1px solid var(--border);
  display: flex; flex-direction: column;
  padding: 0;
  overflow-y: auto;
}

.sidebar-logo {
  padding: 28px 20px 16px;
  border-bottom: 1px solid var(--border);
}
.sidebar-logo h1 {
  font-family: var(--sans); font-size: 17px; font-weight: 800;
  color: white; letter-spacing: -0.3px;
}
.sidebar-logo span { color: var(--accent); font-size: 11px; display: block; margin-top: 3px; letter-spacing: 2px; }

.nav-section-label {
  font-size: 10px; font-weight: 700; letter-spacing: 1.5px;
  color: var(--text-muted); padding: 20px 20px 8px; text-transform: uppercase;
}

.nav-btn {
  display: flex; align-items: center; gap: 10px;
  padding: 11px 20px; cursor: pointer; border: none;
  background: transparent; color: var(--text-muted);
  font-family: var(--mono); font-size: 12.5px; width: 100%;
  text-align: left; transition: all .15s; border-left: 3px solid transparent;
}
.nav-btn:hover { background: #1e2238; color: var(--text); }
.nav-btn.active {
  background: rgba(79,142,247,.08);
  color: var(--accent); border-left-color: var(--accent);
}
.nav-icon { font-size: 15px; width: 20px; text-align: center; }

.sidebar-footer {
  margin-top: auto;
  padding: 20px;
  border-top: 1px solid var(--border);
}
#save-exit-btn {
  width: 100%; padding: 10px; border-radius: var(--radius);
  background: var(--danger); color: white; border: none;
  font-family: var(--mono); font-size: 12px; font-weight: 700;
  cursor: pointer; transition: background .15s;
}
#save-exit-btn:hover { background: #c94040; }

/* ── Main content ───────────────────────────────────────────────────────── */
#main { flex: 1; overflow-y: auto; padding: 32px 36px; }

.page { display: none; }
.page.active { display: block; }

.page-title {
  font-family: var(--sans); font-size: 26px; font-weight: 800;
  color: white; margin-bottom: 24px; letter-spacing: -0.5px;
}

/* ── Cards ──────────────────────────────────────────────────────────────── */
.card {
  background: var(--card); border: 1px solid var(--border);
  border-radius: var(--radius); padding: 24px; margin-bottom: 18px;
}
.card-title {
  font-family: var(--sans); font-size: 15px; font-weight: 700;
  color: white; margin-bottom: 16px;
}

/* ── Stat cards ─────────────────────────────────────────────────────────── */
.stat-row { display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; margin-bottom: 24px; }
.stat-card {
  background: var(--card); border: 1px solid var(--border);
  border-radius: var(--radius); padding: 22px 20px;
  display: flex; flex-direction: column; gap: 4px;
  border-top: 3px solid var(--accent);
}
.stat-card.warn  { border-top-color: var(--warning); }
.stat-card.ok    { border-top-color: var(--success); }
.stat-value { font-family: var(--sans); font-size: 38px; font-weight: 800; color: white; }
.stat-label { font-size: 11px; color: var(--text-muted); letter-spacing: .5px; text-transform: uppercase; }

/* ── Form ───────────────────────────────────────────────────────────────── */
.form-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
.form-group { display: flex; flex-direction: column; gap: 6px; }
.form-group label { font-size: 11px; font-weight: 700; color: var(--text-muted); letter-spacing: .8px; text-transform: uppercase; }
.form-group input, .form-group select {
  background: #0d0f1a; border: 1px solid var(--border);
  color: var(--text); padding: 10px 12px; border-radius: 7px;
  font-family: var(--mono); font-size: 13px; outline: none;
  transition: border-color .15s;
}
.form-group input:focus, .form-group select:focus { border-color: var(--accent); }
.form-group input::placeholder { color: #3a3f60; }

/* ── Buttons ────────────────────────────────────────────────────────────── */
.btn {
  display: inline-flex; align-items: center; gap: 7px;
  padding: 9px 18px; border-radius: 7px; border: none;
  font-family: var(--mono); font-size: 12.5px; font-weight: 700;
  cursor: pointer; transition: all .15s; white-space: nowrap;
}
.btn-primary { background: var(--accent); color: white; }
.btn-primary:hover { background: #3a7de8; }
.btn-success { background: var(--success); color: white; }
.btn-success:hover { background: #2ea66b; }
.btn-danger  { background: var(--danger); color: white; }
.btn-danger:hover  { background: #c94040; }
.btn-ghost {
  background: transparent; color: var(--text-muted);
  border: 1px solid var(--border);
}
.btn-ghost:hover { background: var(--card); color: var(--text); }
.btn-row { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 16px; }

/* ── Toolbar ────────────────────────────────────────────────────────────── */
.toolbar {
  display: flex; align-items: center; gap: 10px;
  margin-bottom: 14px; flex-wrap: wrap;
}
.toolbar input {
  flex: 1; min-width: 240px; max-width: 420px;
  background: var(--card); border: 1px solid var(--border);
  color: var(--text); padding: 9px 14px; border-radius: 7px;
  font-family: var(--mono); font-size: 13px; outline: none;
}
.toolbar input:focus { border-color: var(--accent); }
.toolbar .spacer { flex: 1; }

/* ── Table ──────────────────────────────────────────────────────────────── */
.table-wrap {
  background: var(--card); border: 1px solid var(--border);
  border-radius: var(--radius); overflow: auto;
}
table { width: 100%; border-collapse: collapse; }
thead th {
  background: #0d0f1a; color: var(--text-muted);
  font-size: 10.5px; font-weight: 700; letter-spacing: .8px;
  text-transform: uppercase; padding: 12px 16px; text-align: left;
  white-space: nowrap; border-bottom: 1px solid var(--border);
  position: sticky; top: 0;
}
tbody tr { border-bottom: 1px solid #1e2035; transition: background .1s; cursor: pointer; }
tbody tr:hover { background: #1e2238; }
tbody tr.selected { background: rgba(79,142,247,.12); }
tbody tr:last-child { border-bottom: none; }
td { padding: 11px 16px; color: var(--text); white-space: nowrap; }
td .badge {
  display: inline-block; padding: 2px 8px; border-radius: 20px;
  font-size: 10.5px; font-weight: 700; letter-spacing: .4px;
}
td .badge.pending  { background: rgba(224,154,48,.15); color: var(--warning); }
td .badge.approved { background: rgba(61,186,122,.15); color: var(--success); }
td .badge.denied   { background: rgba(224,82,82,.15);  color: var(--danger);  }

.empty-msg { text-align: center; color: var(--text-muted); padding: 40px; font-size: 13px; }

/* ── Modal ──────────────────────────────────────────────────────────────── */
.modal-overlay {
  position: fixed; inset: 0; background: rgba(0,0,0,.7);
  display: flex; align-items: center; justify-content: center;
  z-index: 100; backdrop-filter: blur(4px);
  opacity: 0; pointer-events: none; transition: opacity .2s;
}
.modal-overlay.open { opacity: 1; pointer-events: all; }
.modal {
  background: var(--card); border: 1px solid var(--border);
  border-radius: 14px; padding: 32px; width: 540px; max-width: 95vw;
  max-height: 90vh; overflow-y: auto;
  transform: translateY(16px); transition: transform .2s;
}
.modal-overlay.open .modal { transform: translateY(0); }
.modal h2 { font-family: var(--sans); font-size: 19px; font-weight: 800; color: white; margin-bottom: 20px; }
.modal-actions { display: flex; gap: 10px; justify-content: flex-end; margin-top: 20px; }

/* ── Toast ──────────────────────────────────────────────────────────────── */
#toast {
  position: fixed; bottom: 28px; right: 28px; z-index: 200;
  background: var(--card); border: 1px solid var(--border);
  border-left: 4px solid var(--accent);
  padding: 14px 20px; border-radius: 8px;
  font-size: 13px; font-weight: 500;
  transform: translateY(80px); opacity: 0;
  transition: all .3s; max-width: 320px;
}
#toast.show { transform: translateY(0); opacity: 1; }
#toast.error   { border-left-color: var(--danger); }
#toast.success { border-left-color: var(--success); }

/* ── Promo inline form ──────────────────────────────────────────────────── */
.inline-form { display: flex; gap: 10px; align-items: flex-end; flex-wrap: wrap; }
.inline-form .form-group { min-width: 160px; }

/* ── Import/Export grid ─────────────────────────────────────────────────── */
.io-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 18px; }
.io-btns { display: flex; flex-wrap: wrap; gap: 10px; margin-top: 4px; }
</style>
</head>
<body>
<div id="app">

  <!-- ── Sidebar ────────────────────────────────────────────────────────── -->
  <nav id="sidebar">
    <div class="sidebar-logo">
      <h1>⚡ User Manager</h1>
      <span>PRO EDITION</span>
    </div>

    <div class="nav-section-label">Navigation</div>

    <button class="nav-btn active" data-tab="dashboard">
      <span class="nav-icon">📊</span> Dashboard
    </button>
    <button class="nav-btn" data-tab="add-user">
      <span class="nav-icon">➕</span> Add User
    </button>
    <button class="nav-btn" data-tab="manage">
      <span class="nav-icon">👥</span> Manage Users
    </button>
    <button class="nav-btn" data-tab="promotions">
      <span class="nav-icon">🏆</span> Promotions
    </button>
    <button class="nav-btn" data-tab="io">
      <span class="nav-icon">📁</span> Import / Export
    </button>

    <div class="sidebar-footer">
      <button id="save-exit-btn">💾 Save state is automatic</button>
    </div>
  </nav>

  <!-- ── Main ──────────────────────────────────────────────────────────── -->
  <main id="main">

    <!-- Dashboard -->
    <section class="page active" id="page-dashboard">
      <div class="page-title">Dashboard</div>
      <div class="stat-row">
        <div class="stat-card">
          <div class="stat-value" id="stat-total">0</div>
          <div class="stat-label">Total Users</div>
        </div>
        <div class="stat-card warn">
          <div class="stat-value" id="stat-pending">0</div>
          <div class="stat-label">Pending Promotions</div>
        </div>
        <div class="stat-card ok">
          <div class="stat-value" id="stat-roles">0</div>
          <div class="stat-label">Unique Roles</div>
        </div>
      </div>
      <div class="card">
        <div class="card-title">All Users</div>
        <div class="table-wrap">
          <table>
            <thead>
              <tr>
                <th>Username</th><th>Full Name</th><th>Email</th>
                <th>Occupation</th><th>Role</th><th>Added</th>
              </tr>
            </thead>
            <tbody id="dash-tbody"></tbody>
          </table>
        </div>
      </div>
    </section>

    <!-- Add User -->
    <section class="page" id="page-add-user">
      <div class="page-title">Add New User</div>
      <div class="card">
        <div class="form-grid">
          <div class="form-group">
            <label>Username *</label>
            <input id="f-username" type="text" placeholder="unique identifier">
          </div>
          <div class="form-group">
            <label>Full Name</label>
            <input id="f-fullname" type="text" placeholder="First Last">
          </div>
          <div class="form-group">
            <label>Email</label>
            <input id="f-email" type="email" placeholder="user@example.com">
          </div>
          <div class="form-group">
            <label>Age</label>
            <input id="f-age" type="text" placeholder="e.g. 28">
          </div>
          <div class="form-group">
            <label>Occupation</label>
            <input id="f-occupation" type="text" placeholder="e.g. Engineer">
          </div>
          <div class="form-group">
            <label>Location</label>
            <input id="f-location" type="text" placeholder="City, Country">
          </div>
          <div class="form-group">
            <label>Role / Level</label>
            <input id="f-role" type="text" placeholder="e.g. Junior, Senior, Manager">
          </div>
        </div>
        <div class="btn-row">
          <button class="btn btn-primary" onclick="addUser()">➕ Add User</button>
          <button class="btn btn-ghost" onclick="clearAddForm()">✕ Clear</button>
        </div>
      </div>
    </section>

    <!-- Manage Users -->
    <section class="page" id="page-manage">
      <div class="page-title">Manage Users</div>
      <div class="toolbar">
        <input id="search-input" type="text" placeholder="🔍  Search username, name, occupation, role…"
               oninput="renderManageTable()">
        <div class="spacer"></div>
        <button class="btn btn-primary" onclick="openEditModal()">✏️ Edit</button>
        <button class="btn btn-danger"  onclick="deleteSelected()">🗑 Delete</button>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>Username</th><th>Full Name</th><th>Email</th>
              <th>Age</th><th>Occupation</th><th>Location</th>
              <th>Role</th><th>Date Added</th>
            </tr>
          </thead>
          <tbody id="manage-tbody"></tbody>
        </table>
      </div>
    </section>

    <!-- Promotions -->
    <section class="page" id="page-promotions">
      <div class="page-title">Promotion Workflow</div>
      <div class="card">
        <div class="card-title">Submit New Request</div>
        <div class="inline-form">
          <div class="form-group">
            <label>Username</label>
            <input id="p-username" type="text" placeholder="exact username">
          </div>
          <div class="form-group">
            <label>Requested Role</label>
            <input id="p-role" type="text" placeholder="e.g. Senior, Manager">
          </div>
          <button class="btn btn-primary" onclick="submitPromo()" style="margin-bottom:1px">Submit</button>
        </div>
      </div>

      <div class="card-title" style="margin-bottom:10px">All Requests</div>
      <div class="toolbar" style="margin-bottom:10px">
        <div class="spacer"></div>
        <button class="btn btn-success" onclick="resolvePromo('approved')">✅ Approve</button>
        <button class="btn btn-danger"  onclick="resolvePromo('denied')">❌ Deny</button>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>Username</th><th>Current Role</th>
              <th>Requested Role</th><th>Date</th><th>Status</th>
            </tr>
          </thead>
          <tbody id="promo-tbody"></tbody>
        </table>
      </div>
    </section>

    <!-- Import / Export -->
    <section class="page" id="page-io">
      <div class="page-title">Import / Export</div>
      <div class="io-grid">
        <div class="card">
          <div class="card-title">📤 Export Users</div>
          <p style="color:var(--text-muted);font-size:12px;margin-bottom:14px">
            Download current user data in your preferred format.
          </p>
          <div class="io-btns">
            <button class="btn btn-primary"  onclick="exportData('json')">JSON</button>
            <button class="btn btn-primary"  onclick="exportData('csv')">CSV</button>
            <button class="btn btn-primary"  onclick="exportData('excel')">Excel</button>
          </div>
        </div>
        <div class="card">
          <div class="card-title">📥 Import Users</div>
          <p style="color:var(--text-muted);font-size:12px;margin-bottom:14px">
            Merge users from a file. Existing usernames will be overwritten.
          </p>
          <div class="io-btns">
            <input type="file" id="import-file" accept=".json,.csv" style="display:none"
                   onchange="doImport(this)">
            <button class="btn btn-ghost" onclick="document.getElementById('import-file').click()">
              📂 Choose File (JSON / CSV)
            </button>
          </div>
        </div>
      </div>
    </section>

  </main>
</div>

<!-- ── Edit Modal ──────────────────────────────────────────────────────────── -->
<div class="modal-overlay" id="edit-modal">
  <div class="modal">
    <h2>✏️ Edit User</h2>
    <input type="hidden" id="edit-username-key">
    <div class="form-grid">
      <div class="form-group">
        <label>Full Name</label>
        <input id="e-fullname" type="text">
      </div>
      <div class="form-group">
        <label>Email</label>
        <input id="e-email" type="email">
      </div>
      <div class="form-group">
        <label>Age</label>
        <input id="e-age" type="text">
      </div>
      <div class="form-group">
        <label>Occupation</label>
        <input id="e-occupation" type="text">
      </div>
      <div class="form-group">
        <label>Location</label>
        <input id="e-location" type="text">
      </div>
      <div class="form-group">
        <label>Role</label>
        <input id="e-role" type="text">
      </div>
    </div>
    <div class="modal-actions">
      <button class="btn btn-ghost"   onclick="closeModal()">Cancel</button>
      <button class="btn btn-primary" onclick="saveEdit()">💾 Save</button>
    </div>
  </div>
</div>

<!-- ── Toast ──────────────────────────────────────────────────────────────── -->
<div id="toast"></div>

<script>
// ── State ──────────────────────────────────────────────────────────────────
let users = {};
let selectedUsername = null;   // manage table
let selectedPromoKey = null;   // promo table  (format: "username::index")

// ── Boot ───────────────────────────────────────────────────────────────────
(async () => {
  await fetchUsers();
  renderDashboard();
})();

// ── Navigation ─────────────────────────────────────────────────────────────
document.querySelectorAll('.nav-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.nav-btn').forEach(b => b.classList.remove('active'));
    document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
    btn.classList.add('active');
    const tab = btn.dataset.tab;
    document.getElementById('page-' + tab).classList.add('active');
    if (tab === 'dashboard')   renderDashboard();
    if (tab === 'manage')      renderManageTable();
    if (tab === 'promotions')  renderPromoTable();
  });
});

// ── API helpers ────────────────────────────────────────────────────────────
async function fetchUsers() {
  const res = await fetch('/api/users');
  users = await res.json();
}

function toast(msg, type = 'success') {
  const el = document.getElementById('toast');
  el.textContent = msg;
  el.className = 'show ' + type;
  setTimeout(() => el.className = '', 3000);
}

// ── Dashboard ──────────────────────────────────────────────────────────────
function renderDashboard() {
  const list = Object.entries(users);
  const pending = list.reduce((n, [, d]) =>
    n + (d.promotion_requests || []).filter(r => r.status === 'pending').length, 0);
  const roles = new Set(list.map(([, d]) => d.role).filter(Boolean)).size;

  document.getElementById('stat-total').textContent   = list.length;
  document.getElementById('stat-pending').textContent = pending;
  document.getElementById('stat-roles').textContent   = roles;

  const tbody = document.getElementById('dash-tbody');
  if (!list.length) {
    tbody.innerHTML = '<tr><td colspan="6" class="empty-msg">No users yet.</td></tr>';
    return;
  }
  tbody.innerHTML = list.map(([u, d]) => `
    <tr>
      <td><strong>${esc(u)}</strong></td>
      <td>${esc(d.full_name||'')}</td>
      <td>${esc(d.email||'')}</td>
      <td>${esc(d.occupation||'')}</td>
      <td>${esc(d.role||'')}</td>
      <td>${esc(d.date_added||'')}</td>
    </tr>`).join('');
}

// ── Add User ───────────────────────────────────────────────────────────────
async function addUser() {
  const username = val('f-username');
  if (!username) { toast('Username is required.', 'error'); return; }

  const payload = {
    full_name:  val('f-fullname'),
    email:      val('f-email'),
    age:        val('f-age'),
    occupation: val('f-occupation'),
    location:   val('f-location'),
    role:       val('f-role') || 'Junior',
  };

  const res = await fetch(`/api/users/${encodeURIComponent(username)}`, {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify(payload),
  });
  const data = await res.json();
  if (!res.ok) { toast(data.error, 'error'); return; }

  users[username] = data.user;
  clearAddForm();
  toast(`User '${username}' added.`, 'success');
}

function clearAddForm() {
  ['f-username','f-fullname','f-email','f-age','f-occupation','f-location','f-role']
    .forEach(id => document.getElementById(id).value = '');
}

// ── Manage Users ───────────────────────────────────────────────────────────
function renderManageTable() {
  const q = (document.getElementById('search-input')?.value || '').toLowerCase();
  const tbody = document.getElementById('manage-tbody');
  selectedUsername = null;

  const filtered = Object.entries(users).filter(([u, d]) => !q || [
    u, d.full_name, d.email, d.occupation, d.location, d.role
  ].some(v => (v||'').toLowerCase().includes(q)));

  if (!filtered.length) {
    tbody.innerHTML = '<tr><td colspan="8" class="empty-msg">No users found.</td></tr>';
    return;
  }
  tbody.innerHTML = filtered.map(([u, d]) => `
    <tr data-username="${esc(u)}" onclick="selectUser('${esc(u)}', this)">
      <td><strong>${esc(u)}</strong></td>
      <td>${esc(d.full_name||'')}</td>
      <td>${esc(d.email||'')}</td>
      <td>${esc(d.age||'')}</td>
      <td>${esc(d.occupation||'')}</td>
      <td>${esc(d.location||'')}</td>
      <td>${esc(d.role||'')}</td>
      <td>${esc(d.date_added||'')}</td>
    </tr>`).join('');
}

function selectUser(username, row) {
  document.querySelectorAll('#manage-tbody tr').forEach(r => r.classList.remove('selected'));
  row.classList.add('selected');
  selectedUsername = username;
}

function openEditModal() {
  if (!selectedUsername) { toast('Select a user first.', 'error'); return; }
  const d = users[selectedUsername];
  document.getElementById('edit-username-key').value = selectedUsername;
  document.getElementById('e-fullname').value   = d.full_name   || '';
  document.getElementById('e-email').value      = d.email       || '';
  document.getElementById('e-age').value        = d.age         || '';
  document.getElementById('e-occupation').value = d.occupation  || '';
  document.getElementById('e-location').value   = d.location    || '';
  document.getElementById('e-role').value       = d.role        || '';
  document.getElementById('edit-modal').classList.add('open');
}

function closeModal() {
  document.getElementById('edit-modal').classList.remove('open');
}

async function saveEdit() {
  const uname = document.getElementById('edit-username-key').value;
  const payload = {
    full_name:  document.getElementById('e-fullname').value.trim(),
    email:      document.getElementById('e-email').value.trim(),
    age:        document.getElementById('e-age').value.trim(),
    occupation: document.getElementById('e-occupation').value.trim(),
    location:   document.getElementById('e-location').value.trim(),
    role:       document.getElementById('e-role').value.trim(),
  };

  const res = await fetch(`/api/users/${encodeURIComponent(uname)}`, {
    method: 'PUT',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify(payload),
  });
  if (!res.ok) { const d = await res.json(); toast(d.error, 'error'); return; }
  const data = await res.json();
  users[uname] = { ...users[uname], ...data.user };
  closeModal();
  renderManageTable();
  toast(`User '${uname}' updated.`);
}

async function deleteSelected() {
  if (!selectedUsername) { toast('Select a user first.', 'error'); return; }
  if (!confirm(`Delete user '${selectedUsername}'? This cannot be undone.`)) return;
  const res = await fetch(`/api/users/${encodeURIComponent(selectedUsername)}`, { method: 'DELETE' });
  if (!res.ok) { const d = await res.json(); toast(d.error, 'error'); return; }
  delete users[selectedUsername];
  selectedUsername = null;
  renderManageTable();
  toast('User deleted.');
}

// ── Promotions ─────────────────────────────────────────────────────────────
function renderPromoTable() {
  const tbody = document.getElementById('promo-tbody');
  selectedPromoKey = null;
  const rows = [];
  for (const [uname, d] of Object.entries(users)) {
    (d.promotion_requests || []).forEach((req, i) => {
      rows.push({ uname, i, req });
    });
  }
  if (!rows.length) {
    tbody.innerHTML = '<tr><td colspan="5" class="empty-msg">No promotion requests yet.</td></tr>';
    return;
  }
  tbody.innerHTML = rows.map(({ uname, i, req }) => `
    <tr data-key="${esc(uname)}::${i}"
        onclick="selectPromo('${esc(uname)}::${i}', this)">
      <td><strong>${esc(uname)}</strong></td>
      <td>${esc(req.current_role||'')}</td>
      <td>${esc(req.requested_role||'')}</td>
      <td>${esc(req.date||'')}</td>
      <td><span class="badge ${req.status}">${req.status}</span></td>
    </tr>`).join('');
}

function selectPromo(key, row) {
  document.querySelectorAll('#promo-tbody tr').forEach(r => r.classList.remove('selected'));
  row.classList.add('selected');
  selectedPromoKey = key;
}

async function submitPromo() {
  const uname = val('p-username');
  const role  = val('p-role');
  if (!uname || !role) { toast('Username and requested role are required.', 'error'); return; }

  const res = await fetch('/api/promotions', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({ username: uname, requested_role: role }),
  });
  const data = await res.json();
  if (!res.ok) { toast(data.error, 'error'); return; }

  users[uname] = data.user;
  document.getElementById('p-username').value = '';
  document.getElementById('p-role').value = '';
  renderPromoTable();
  toast(`Request submitted for '${uname}'.`);
}

async function resolvePromo(resolution) {
  if (!selectedPromoKey) { toast('Select a promotion request first.', 'error'); return; }
  const [uname, idx] = selectedPromoKey.split('::');

  const res = await fetch(`/api/promotions/${encodeURIComponent(uname)}/${idx}`, {
    method: 'PUT',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({ resolution }),
  });
  const data = await res.json();
  if (!res.ok) { toast(data.error, 'error'); return; }

  users[uname] = data.user;
  renderPromoTable();
  toast(`Request ${resolution} for '${uname}'.`, resolution === 'approved' ? 'success' : 'error');
}

// ── Import / Export ─────────────────────────────────────────────────────────
function exportData(fmt) {
  window.location.href = `/api/export/${fmt}`;
}

async function doImport(input) {
  const file = input.files[0];
  if (!file) return;
  const formData = new FormData();
  formData.append('file', file);
  const res = await fetch('/api/import', { method: 'POST', body: formData });
  const data = await res.json();
  if (!res.ok) { toast(data.error, 'error'); return; }
  await fetchUsers();
  renderDashboard();
  toast(`Imported ${data.count} users from ${file.name}.`);
  input.value = '';
}

// ── Utilities ───────────────────────────────────────────────────────────────
function val(id) { return document.getElementById(id).value.trim(); }
function esc(s)  { return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }

// Close modal on overlay click
document.getElementById('edit-modal').addEventListener('click', function(e) {
  if (e.target === this) closeModal();
});
</script>
</body>
</html>"""


# ═════════════════════════════════════════════════════════════════════════════
#  API ROUTES
# ═════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template_string(HTML)


# ── Users ─────────────────────────────────────────────────────────────────────
@app.route("/api/users", methods=["GET"])
def get_users():
    return jsonify(load_users())


@app.route("/api/users/<username>", methods=["POST"])
def add_user(username: str):
    users = load_users()
    if username in users:
        return jsonify({"error": f"User '{username}' already exists."}), 409

    data = request.get_json() or {}
    user = {
        "full_name":          data.get("full_name", ""),
        "email":              data.get("email", ""),
        "age":                data.get("age", ""),
        "occupation":         data.get("occupation", ""),
        "location":           data.get("location", ""),
        "role":               data.get("role", "Junior"),
        "promotion_requests": [],
        "date_added":         datetime.now().strftime("%Y-%m-%d"),
    }
    users[username] = user
    save_users(users)
    return jsonify({"user": user}), 201


@app.route("/api/users/<username>", methods=["PUT"])
def update_user(username: str):
    users = load_users()
    if username not in users:
        return jsonify({"error": f"User '{username}' not found."}), 404

    data = request.get_json() or {}
    editable = ["full_name", "email", "age", "occupation", "location", "role"]
    for key in editable:
        if key in data:
            users[username][key] = data[key]
    save_users(users)
    return jsonify({"user": users[username]})


@app.route("/api/users/<username>", methods=["DELETE"])
def delete_user(username: str):
    users = load_users()
    if username not in users:
        return jsonify({"error": f"User '{username}' not found."}), 404
    del users[username]
    save_users(users)
    return jsonify({"message": "deleted"})


# ── Promotions ────────────────────────────────────────────────────────────────
@app.route("/api/promotions", methods=["POST"])
def submit_promotion():
    users = load_users()
    data  = request.get_json() or {}
    uname = data.get("username", "").strip()
    role  = data.get("requested_role", "").strip()

    if uname not in users:
        return jsonify({"error": f"User '{uname}' not found."}), 404
    if not role:
        return jsonify({"error": "requested_role is required."}), 400

    req = {
        "requested_role": role,
        "current_role":   users[uname].get("role", ""),
        "status":         "pending",
        "date":           datetime.now().strftime("%Y-%m-%d"),
        "resolved_date":  "",
    }
    users[uname].setdefault("promotion_requests", []).append(req)
    save_users(users)
    return jsonify({"user": users[uname]}), 201


@app.route("/api/promotions/<username>/<int:idx>", methods=["PUT"])
def resolve_promotion(username: str, idx: int):
    users = load_users()
    if username not in users:
        return jsonify({"error": f"User '{username}' not found."}), 404

    reqs = users[username].get("promotion_requests", [])
    if idx >= len(reqs):
        return jsonify({"error": "Promotion request not found."}), 404

    req = reqs[idx]
    if req["status"] != "pending":
        return jsonify({"error": "Request already resolved."}), 409

    resolution = (request.get_json() or {}).get("resolution", "")
    if resolution not in ("approved", "denied"):
        return jsonify({"error": "resolution must be 'approved' or 'denied'."}), 400

    req["status"]        = resolution
    req["resolved_date"] = datetime.now().strftime("%Y-%m-%d")
    if resolution == "approved":
        users[username]["role"] = req["requested_role"]

    save_users(users)
    return jsonify({"user": users[username]})


# ── Export ────────────────────────────────────────────────────────────────────
@app.route("/api/export/json")
def export_json():
    users = load_users()
    buf   = io.BytesIO(json.dumps(users, indent=4).encode())
    buf.seek(0)
    return send_file(buf, mimetype="application/json",
                     as_attachment=True, download_name="users.json")


@app.route("/api/export/csv")
def export_csv():
    users   = load_users()
    headers = ["username", "full_name", "email", "age",
               "occupation", "location", "role", "date_added"]
    buf = io.StringIO()
    w   = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    w.writeheader()
    for uname, d in users.items():
        row = {"username": uname}
        row.update({k: d.get(k, "") for k in headers[1:]})
        w.writerow(row)
    buf.seek(0)
    return send_file(
        io.BytesIO(buf.getvalue().encode()),
        mimetype="text/csv",
        as_attachment=True,
        download_name="users.csv",
    )


@app.route("/api/export/excel")
def export_excel():
    if not EXCEL_OK:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    users = load_users()
    wb    = openpyxl.Workbook()
    ws    = wb.active
    ws.title = "Users"

    hdr_fill = PatternFill("solid", fgColor="0d1a40")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    headers  = ["Username", "Full Name", "Email", "Age",
                "Occupation", "Location", "Role", "Date Added"]
    keys     = ["", "full_name", "email", "age",
                "occupation", "location", "role", "date_added"]

    for col_i, h in enumerate(headers, 1):
        cell       = ws.cell(row=1, column=col_i, value=h)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for row_i, (uname, d) in enumerate(users.items(), start=2):
        ws.cell(row=row_i, column=1, value=uname)
        for col_i, key in enumerate(keys[1:], 2):
            ws.cell(row=row_i, column=col_i, value=d.get(key, ""))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="users.xlsx")


# ── Import ────────────────────────────────────────────────────────────────────
@app.route("/api/import", methods=["POST"])
def import_users():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded."}), 400

    f    = request.files["file"]
    name = f.filename.lower()
    users = load_users()
    count = 0

    try:
        if name.endswith(".json"):
            raw = json.load(f)
            if isinstance(raw, dict):
                users.update(raw)
                count = len(raw)
            elif isinstance(raw, list):
                for u in raw:
                    uname = u.get("username") or u.get("name")
                    if uname:
                        users[uname] = {k: v for k, v in u.items() if k != "username"}
                        users[uname].setdefault("promotion_requests", [])
                        count += 1
        elif name.endswith(".csv"):
            text   = f.read().decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                uname = row.get("username") or row.get("name")
                if uname:
                    users[uname] = {k: v for k, v in row.items()
                                    if k not in ("username", "name")}
                    users[uname].setdefault("promotion_requests", [])
                    count += 1
        else:
            return jsonify({"error": "Only .json and .csv files are supported."}), 400

        save_users(users)
        return jsonify({"count": count})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Run ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n  ⚡ User Manager Pro — Web Edition")
    print("  ─────────────────────────────────")
    print("  http://localhost:5000\n")
    app.run(debug=True, port=5000)
