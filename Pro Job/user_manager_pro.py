"""
User Manager Pro — CustomTkinter Desktop App
=============================================
Merged & upgraded from user_manager_gui.py + user_collector.py

Features:
  - Dashboard with live stats
  - Add / Edit / Delete users
  - Search & filter
  - Promotion request & approval workflow
  - Import from JSON / CSV
  - Export to JSON / CSV / Excel
  - Dark / Light mode toggle
  - Auto-save on every write operation

Requirements:
  pip install customtkinter openpyxl
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import csv
from datetime import datetime

# ── Optional Excel support ──────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── Config ───────────────────────────────────────────────────────────────────
DATA_FILE = "users.json"
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

ACCENT      = "#4f8ef7"
DANGER      = "#e05252"
SUCCESS     = "#3dba7a"
WARNING     = "#e09a30"
SIDEBAR_BG  = "#1a1d2e"
CARD_BG     = "#232640"
# Fonts are defined as None here; initialised inside UserManagerPro.__init__
# after the Tk root window exists (CTkFont requires a live root).
BODY_FONT   = None
HEADING     = None
SUBHEADING  = None


# ── Data helpers ─────────────────────────────────────────────────────────────
def load_users() -> dict:
    if not os.path.exists(DATA_FILE):
        return {}
    try:
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return {}


def save_users(users: dict) -> None:
    with open(DATA_FILE, "w") as f:
        json.dump(users, f, indent=4)


# ── Edit-user dialog ──────────────────────────────────────────────────────────
class EditUserDialog(ctk.CTkToplevel):
    FIELDS = ["full_name", "email", "age", "occupation", "location", "role"]

    def __init__(self, parent, username: str, user_data: dict, on_save):
        super().__init__(parent)
        self.title(f"Edit — {username}")
        self.geometry("480x460")
        self.resizable(False, False)
        self.grab_set()

        self.username = username
        self.data = dict(user_data)
        self.on_save = on_save

        ctk.CTkLabel(self, text=f"✏️  Editing: {username}",
                     font=ctk.CTkFont(size=17, weight="bold")).pack(pady=(22, 12))

        scroller = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroller.pack(fill="both", expand=True, padx=24, pady=4)
        scroller.columnconfigure(1, weight=1)

        self.vars: dict[str, tk.StringVar] = {}
        for i, key in enumerate(self.FIELDS):
            label = key.replace("_", " ").title()
            ctk.CTkLabel(scroller, text=label,
                         font=ctk.CTkFont(size=12, weight="bold")).grid(
                row=i, column=0, padx=8, pady=7, sticky="w")
            var = tk.StringVar(value=self.data.get(key, ""))
            ctk.CTkEntry(scroller, textvariable=var, height=34).grid(
                row=i, column=1, padx=8, pady=7, sticky="ew")
            self.vars[key] = var

        ctk.CTkButton(self, text="💾  Save Changes", height=42,
                      fg_color=ACCENT, hover_color="#3a7de8",
                      font=ctk.CTkFont(size=13, weight="bold"),
                      command=self._save).pack(pady=18)

    def _save(self):
        for key, var in self.vars.items():
            self.data[key] = var.get().strip()
        self.on_save(self.username, self.data)
        self.destroy()


# ── Main app ──────────────────────────────────────────────────────────────────
class UserManagerPro(ctk.CTk):

    def __init__(self):
        super().__init__()
        self.title("User Manager Pro")
        self.geometry("1240x800")
        self.minsize(960, 620)

        # CTkFont requires a live root window — initialise here, not at module level
        global BODY_FONT, HEADING, SUBHEADING
        BODY_FONT  = ctk.CTkFont(family="Segoe UI", size=13)
        HEADING    = ctk.CTkFont(family="Segoe UI", size=26, weight="bold")
        SUBHEADING = ctk.CTkFont(family="Segoe UI", size=16, weight="bold")

        self.users = load_users()
        self._build_ui()

    # ── Layout skeleton ──────────────────────────────────────────────────────
    def _build_ui(self):
        # ── Sidebar ──────────────────────────────────────────────────────────
        self.sidebar = ctk.CTkFrame(self, width=230, corner_radius=0,
                                    fg_color=SIDEBAR_BG)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        ctk.CTkLabel(self.sidebar, text="⚡ User Manager",
                     font=ctk.CTkFont(size=19, weight="bold"),
                     text_color="white").pack(pady=(32, 6), padx=20)
        ctk.CTkLabel(self.sidebar, text="PRO EDITION",
                     font=ctk.CTkFont(size=9, weight="bold"),
                     text_color=ACCENT).pack(padx=20, pady=(0, 24))

        ctk.CTkLabel(self.sidebar, text="NAVIGATION",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     text_color="gray50").pack(padx=20, pady=(0, 6), anchor="w")

        self.nav_btns: dict[str, ctk.CTkButton] = {}
        nav_items = [
            ("📊", "Dashboard"),
            ("➕", "Add User"),
            ("👥", "Manage Users"),
            ("🏆", "Promotions"),
            ("📁", "Import / Export"),
        ]
        for icon, key in nav_items:
            btn = ctk.CTkButton(
                self.sidebar, text=f"  {icon}  {key}", width=210, height=40,
                anchor="w", fg_color="transparent",
                text_color=("gray90", "gray85"),
                hover_color="#2a2f4a",
                font=ctk.CTkFont(size=13),
                command=lambda k=key: self.show_tab(k),
            )
            btn.pack(pady=2, padx=10)
            self.nav_btns[key] = btn

        # Settings section
        ctk.CTkLabel(self.sidebar, text="SETTINGS",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     text_color="gray50").pack(padx=20, pady=(28, 6), anchor="w")
        self.theme_sw = ctk.CTkSwitch(
            self.sidebar, text="Dark Mode",
            font=ctk.CTkFont(size=12),
            command=self._toggle_theme,
        )
        self.theme_sw.pack(padx=24, anchor="w")
        self.theme_sw.select()

        ctk.CTkButton(
            self.sidebar, text="💾  Save & Exit", width=210, height=42,
            fg_color=DANGER, hover_color="#c94040",
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._save_and_exit,
        ).pack(side="bottom", pady=24, padx=10)

        # ── Content area ─────────────────────────────────────────────────────
        self.content = ctk.CTkFrame(self, fg_color="transparent")
        self.content.pack(side="right", fill="both", expand=True, padx=18, pady=18)

        self.tab_frames: dict[str, ctk.CTkFrame] = {}
        self._build_dashboard()
        self._build_add_user()
        self._build_manage_users()
        self._build_promotions()
        self._build_import_export()

        self.show_tab("Dashboard")

    # ── Navigation ───────────────────────────────────────────────────────────
    def show_tab(self, name: str):
        for frame in self.tab_frames.values():
            frame.pack_forget()
        self.tab_frames[name].pack(fill="both", expand=True)
        for k, btn in self.nav_btns.items():
            btn.configure(fg_color=("#1e3a6e", "#1f3d7a") if k == name else "transparent")
        if name == "Dashboard":
            self._refresh_dashboard()
        elif name == "Manage Users":
            self._refresh_user_tree()
        elif name == "Promotions":
            self._refresh_promotions()

    def _toggle_theme(self):
        ctk.set_appearance_mode("dark" if self.theme_sw.get() else "light")

    def _save_and_exit(self):
        save_users(self.users)
        self.destroy()

    # ═══════════════════════════════════════════════════════════════════════
    #  TAB 1 — Dashboard
    # ═══════════════════════════════════════════════════════════════════════
    def _build_dashboard(self):
        f = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tab_frames["Dashboard"] = f

        ctk.CTkLabel(f, text="Dashboard", font=HEADING).pack(anchor="w", pady=(0, 20))

        self.stat_row = ctk.CTkFrame(f, fg_color="transparent")
        self.stat_row.pack(fill="x")

        ctk.CTkLabel(f, text="All Users", font=SUBHEADING).pack(anchor="w", pady=(24, 6))
        self.dash_box = ctk.CTkTextbox(f, font=ctk.CTkFont(family="Courier New", size=12),
                                       state="disabled")
        self.dash_box.pack(fill="both", expand=True)

    def _refresh_dashboard(self):
        for w in self.stat_row.winfo_children():
            w.destroy()

        total = len(self.users)
        pending = sum(
            1 for u in self.users.values()
            if any(r.get("status") == "pending"
                   for r in u.get("promotion_requests", []))
        )
        roles = len({u.get("role", "") for u in self.users.values() if u.get("role")})

        for label, value, color in [
            ("Total Users",         total,   ACCENT),
            ("Pending Promotions",  pending, WARNING),
            ("Unique Roles",        roles,   SUCCESS),
        ]:
            card = ctk.CTkFrame(self.stat_row, fg_color=color, corner_radius=14)
            card.pack(side="left", padx=8, pady=4, fill="x", expand=True)
            ctk.CTkLabel(card, text=str(value),
                         font=ctk.CTkFont(size=40, weight="bold"),
                         text_color="white").pack(pady=(18, 2))
            ctk.CTkLabel(card, text=label,
                         font=ctk.CTkFont(size=12),
                         text_color="white").pack(pady=(0, 18))

        self.dash_box.configure(state="normal")
        self.dash_box.delete("1.0", "end")
        if not self.users:
            self.dash_box.insert("end", "  No users yet — add one to get started!\n")
        else:
            header = f"  {'USERNAME':<20} {'FULL NAME':<24} {'ROLE':<14} {'OCCUPATION'}\n"
            self.dash_box.insert("end", header)
            self.dash_box.insert("end", "  " + "─" * 72 + "\n")
            for uname, d in self.users.items():
                line = (f"  {uname:<20} {d.get('full_name',''):<24}"
                        f" {d.get('role',''):<14} {d.get('occupation','')}\n")
                self.dash_box.insert("end", line)
        self.dash_box.configure(state="disabled")

    # ═══════════════════════════════════════════════════════════════════════
    #  TAB 2 — Add User
    # ═══════════════════════════════════════════════════════════════════════
    def _build_add_user(self):
        f = ctk.CTkScrollableFrame(self.content, fg_color="transparent")
        self.tab_frames["Add User"] = f

        ctk.CTkLabel(f, text="Add New User", font=HEADING).pack(anchor="w", pady=(0, 20))

        card = ctk.CTkFrame(f, fg_color=CARD_BG, corner_radius=14)
        card.pack(fill="x", padx=4, pady=6)
        card.columnconfigure(1, weight=1)

        self.add_vars: dict[str, tk.StringVar] = {}
        fields = [
            ("Username *",   "username",   "Required — must be unique"),
            ("Full Name",    "full_name",  ""),
            ("Email",        "email",      ""),
            ("Age",          "age",        ""),
            ("Occupation",   "occupation", ""),
            ("Location",     "location",   "City, State, Country"),
            ("Role / Level", "role",       "e.g. Junior, Senior, Manager"),
        ]
        for i, (label, key, hint) in enumerate(fields):
            ctk.CTkLabel(card, text=label,
                         font=ctk.CTkFont(size=12, weight="bold")).grid(
                row=i, column=0, padx=24, pady=9, sticky="w")
            var = tk.StringVar()
            ent = ctk.CTkEntry(card, textvariable=var, height=36,
                               placeholder_text=hint, font=BODY_FONT)
            ent.grid(row=i, column=1, padx=24, pady=9, sticky="ew")
            self.add_vars[key] = var

        ctk.CTkButton(
            f, text="➕  Add User", height=44,
            fg_color=ACCENT, hover_color="#3a7de8",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._add_user,
        ).pack(pady=20)

    def _add_user(self):
        username = self.add_vars["username"].get().strip()
        if not username:
            messagebox.showerror("Missing Field", "Username is required.")
            return
        if username in self.users:
            messagebox.showerror("Duplicate", f"'{username}' already exists.")
            return

        self.users[username] = {
            "full_name":          self.add_vars["full_name"].get().strip(),
            "email":              self.add_vars["email"].get().strip(),
            "age":                self.add_vars["age"].get().strip(),
            "occupation":         self.add_vars["occupation"].get().strip(),
            "location":           self.add_vars["location"].get().strip(),
            "role":               self.add_vars["role"].get().strip() or "Junior",
            "promotion_requests": [],
            "date_added":         datetime.now().strftime("%Y-%m-%d"),
        }

        for var in self.add_vars.values():
            var.set("")

        save_users(self.users)
        messagebox.showinfo("✅ Added", f"User '{username}' added successfully.")

    # ═══════════════════════════════════════════════════════════════════════
    #  TAB 3 — Manage Users
    # ═══════════════════════════════════════════════════════════════════════
    def _build_manage_users(self):
        f = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tab_frames["Manage Users"] = f

        ctk.CTkLabel(f, text="Manage Users", font=HEADING).pack(anchor="w", pady=(0, 12))

        # ── Toolbar ──────────────────────────────────────────────────────────
        toolbar = ctk.CTkFrame(f, fg_color="transparent")
        toolbar.pack(fill="x", pady=(0, 10))

        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self._refresh_user_tree())
        ctk.CTkEntry(
            toolbar, textvariable=self.search_var,
            placeholder_text="🔍  Search by username, name, occupation…",
            width=380, height=38, font=BODY_FONT,
        ).pack(side="left")

        ctk.CTkButton(toolbar, text="🗑  Delete",
                      fg_color=DANGER, hover_color="#c94040", height=38, width=110,
                      command=self._delete_user).pack(side="right", padx=4)
        ctk.CTkButton(toolbar, text="✏️  Edit",
                      fg_color=ACCENT, hover_color="#3a7de8", height=38, width=110,
                      command=self._edit_user).pack(side="right", padx=4)

        # ── Treeview ─────────────────────────────────────────────────────────
        tree_card = ctk.CTkFrame(f, corner_radius=14)
        tree_card.pack(fill="both", expand=True)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Pro.Treeview",
                        background="#1e2235", foreground="#e0e4f7",
                        fieldbackground="#1e2235", rowheight=30,
                        font=("Segoe UI", 11))
        style.configure("Pro.Treeview.Heading",
                        background="#151829", foreground="#a0aacc",
                        font=("Segoe UI", 11, "bold"), relief="flat")
        style.map("Pro.Treeview",
                  background=[("selected", "#2d4080")],
                  foreground=[("selected", "white")])

        cols = ("Username", "Full Name", "Email", "Age", "Occupation", "Location", "Role", "Added")
        self.tree = ttk.Treeview(tree_card, columns=cols,
                                 show="headings", selectmode="browse",
                                 style="Pro.Treeview")

        col_widths = {"Username": 130, "Full Name": 160, "Email": 190,
                      "Age": 55, "Occupation": 140, "Location": 140,
                      "Role": 100, "Added": 100}
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths.get(col, 120), minwidth=50)

        vsb = ttk.Scrollbar(tree_card, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_card, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_card.rowconfigure(0, weight=1)
        tree_card.columnconfigure(0, weight=1)

    def _refresh_user_tree(self):
        query = getattr(self, "search_var", None)
        q = query.get().lower().strip() if query else ""
        for row in self.tree.get_children():
            self.tree.delete(row)
        for uname, d in self.users.items():
            if q and not any(q in str(v).lower() for v in [
                uname, d.get("full_name",""), d.get("occupation",""),
                d.get("email",""), d.get("location",""), d.get("role","")
            ]):
                continue
            self.tree.insert("", "end", iid=uname, values=(
                uname,
                d.get("full_name", ""),
                d.get("email", ""),
                d.get("age", ""),
                d.get("occupation", ""),
                d.get("location", ""),
                d.get("role", ""),
                d.get("date_added", ""),
            ))

    def _selected_username(self) -> str | None:
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("No Selection", "Select a user from the table first.")
            return None
        return sel[0]

    def _delete_user(self):
        uname = self._selected_username()
        if not uname:
            return
        if messagebox.askyesno("Confirm Delete",
                               f"Permanently delete user '{uname}'?\nThis cannot be undone."):
            del self.users[uname]
            save_users(self.users)
            self._refresh_user_tree()
            messagebox.showinfo("Deleted", f"User '{uname}' has been removed.")

    def _edit_user(self):
        uname = self._selected_username()
        if not uname:
            return
        EditUserDialog(self, uname, self.users[uname], self._on_edit_save)

    def _on_edit_save(self, uname: str, updated: dict):
        self.users[uname] = updated
        save_users(self.users)
        self._refresh_user_tree()
        messagebox.showinfo("Updated", f"User '{uname}' updated.")

    # ═══════════════════════════════════════════════════════════════════════
    #  TAB 4 — Promotions
    # ═══════════════════════════════════════════════════════════════════════
    def _build_promotions(self):
        f = ctk.CTkFrame(self.content, fg_color="transparent")
        self.tab_frames["Promotions"] = f

        ctk.CTkLabel(f, text="Promotion Workflow", font=HEADING).pack(anchor="w", pady=(0, 14))

        # ── Submit request ────────────────────────────────────────────────────
        req_card = ctk.CTkFrame(f, fg_color=CARD_BG, corner_radius=14)
        req_card.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(req_card, text="Submit New Request",
                     font=SUBHEADING).pack(anchor="w", padx=22, pady=(18, 10))

        row = ctk.CTkFrame(req_card, fg_color="transparent")
        row.pack(fill="x", padx=22, pady=(0, 18))

        self.promo_uname_var = tk.StringVar()
        self.promo_role_var  = tk.StringVar()

        ctk.CTkLabel(row, text="Username:", font=BODY_FONT).pack(side="left")
        ctk.CTkEntry(row, textvariable=self.promo_uname_var,
                     width=190, height=36, font=BODY_FONT,
                     placeholder_text="exact username").pack(side="left", padx=(6, 18))
        ctk.CTkLabel(row, text="Requested Role:", font=BODY_FONT).pack(side="left")
        ctk.CTkEntry(row, textvariable=self.promo_role_var,
                     width=190, height=36, font=BODY_FONT,
                     placeholder_text="e.g. Senior, Manager").pack(side="left", padx=(6, 18))
        ctk.CTkButton(row, text="Submit", height=36, width=110,
                      fg_color=ACCENT, hover_color="#3a7de8",
                      command=self._submit_promo).pack(side="left")

        # ── Requests table ────────────────────────────────────────────────────
        ctk.CTkLabel(f, text="All Requests", font=SUBHEADING).pack(anchor="w", pady=(0, 6))

        tbl_card = ctk.CTkFrame(f, corner_radius=14)
        tbl_card.pack(fill="both", expand=True)

        pcols = ("Username", "Current Role", "Requested Role", "Date", "Status")
        self.promo_tree = ttk.Treeview(tbl_card, columns=pcols,
                                       show="headings", selectmode="browse",
                                       style="Pro.Treeview")
        for col in pcols:
            self.promo_tree.heading(col, text=col)
            self.promo_tree.column(col, width=160, minwidth=80)

        vsb = ttk.Scrollbar(tbl_card, orient="vertical", command=self.promo_tree.yview)
        self.promo_tree.configure(yscrollcommand=vsb.set)
        self.promo_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tbl_card.rowconfigure(0, weight=1)
        tbl_card.columnconfigure(0, weight=1)

        # ── Approve / Deny buttons ────────────────────────────────────────────
        action_row = ctk.CTkFrame(f, fg_color="transparent")
        action_row.pack(fill="x", pady=12)
        ctk.CTkButton(action_row, text="✅  Approve",
                      fg_color=SUCCESS, hover_color="#2ea66b", height=40, width=140,
                      command=lambda: self._resolve_promo("approved")).pack(side="left", padx=4)
        ctk.CTkButton(action_row, text="❌  Deny",
                      fg_color=DANGER, hover_color="#c94040", height=40, width=140,
                      command=lambda: self._resolve_promo("denied")).pack(side="left", padx=4)

    def _submit_promo(self):
        uname = self.promo_uname_var.get().strip()
        role  = self.promo_role_var.get().strip()
        if uname not in self.users:
            messagebox.showerror("Not Found", f"User '{uname}' not found.")
            return
        if not role:
            messagebox.showerror("Missing", "Requested role is required.")
            return
        req = {
            "requested_role": role,
            "current_role":   self.users[uname].get("role", ""),
            "status":         "pending",
            "date":           datetime.now().strftime("%Y-%m-%d"),
            "resolved_date":  "",
        }
        self.users[uname].setdefault("promotion_requests", []).append(req)
        save_users(self.users)
        self.promo_uname_var.set("")
        self.promo_role_var.set("")
        self._refresh_promotions()
        messagebox.showinfo("Submitted", f"Promotion request submitted for '{uname}'.")

    def _refresh_promotions(self):
        for row in self.promo_tree.get_children():
            self.promo_tree.delete(row)
        for uname, data in self.users.items():
            for i, req in enumerate(data.get("promotion_requests", [])):
                iid = f"{uname}::{i}"
                self.promo_tree.insert("", "end", iid=iid, values=(
                    uname,
                    req.get("current_role", ""),
                    req.get("requested_role", ""),
                    req.get("date", ""),
                    req.get("status", "").capitalize(),
                ))

    def _resolve_promo(self, resolution: str):
        sel = self.promo_tree.selection()
        if not sel:
            messagebox.showwarning("No Selection", "Select a promotion request first.")
            return
        uname, idx_s = sel[0].split("::")
        req = self.users[uname]["promotion_requests"][int(idx_s)]
        if req["status"] != "pending":
            messagebox.showwarning("Already Resolved",
                                   "This request has already been resolved.")
            return
        req["status"]        = resolution
        req["resolved_date"] = datetime.now().strftime("%Y-%m-%d")
        if resolution == "approved":
            self.users[uname]["role"] = req["requested_role"]
        save_users(self.users)
        self._refresh_promotions()
        label = "approved ✅" if resolution == "approved" else "denied ❌"
        messagebox.showinfo("Done", f"Promotion request {label} for '{uname}'.")

    # ═══════════════════════════════════════════════════════════════════════
    #  TAB 5 — Import / Export
    # ═══════════════════════════════════════════════════════════════════════
    def _build_import_export(self):
        f = ctk.CTkScrollableFrame(self.content, fg_color="transparent")
        self.tab_frames["Import / Export"] = f

        ctk.CTkLabel(f, text="Import / Export", font=HEADING).pack(anchor="w", pady=(0, 20))

        # Export card
        exp = ctk.CTkFrame(f, fg_color=CARD_BG, corner_radius=14)
        exp.pack(fill="x", padx=4, pady=8)
        ctk.CTkLabel(exp, text="📤  Export Users",
                     font=SUBHEADING).pack(anchor="w", padx=22, pady=(18, 10))
        r = ctk.CTkFrame(exp, fg_color="transparent")
        r.pack(padx=22, pady=(0, 20), anchor="w")
        for label, cmd in [
            ("Export JSON",  self._export_json),
            ("Export CSV",   self._export_csv),
            ("Export Excel", self._export_excel),
        ]:
            ctk.CTkButton(r, text=label, width=150, height=38,
                          fg_color=ACCENT, hover_color="#3a7de8",
                          command=cmd).pack(side="left", padx=6)

        # Import card
        imp = ctk.CTkFrame(f, fg_color=CARD_BG, corner_radius=14)
        imp.pack(fill="x", padx=4, pady=8)
        ctk.CTkLabel(imp, text="📥  Import Users",
                     font=SUBHEADING).pack(anchor="w", padx=22, pady=(18, 10))
        r2 = ctk.CTkFrame(imp, fg_color="transparent")
        r2.pack(padx=22, pady=(0, 20), anchor="w")
        for label, cmd in [
            ("Import JSON", self._import_json),
            ("Import CSV",  self._import_csv),
        ]:
            ctk.CTkButton(r2, text=label, width=150, height=38,
                          fg_color="#555", hover_color="#666",
                          command=cmd).pack(side="left", padx=6)

        ctk.CTkLabel(f,
                     text="ℹ️  Import merges new users with existing data. "
                          "Existing usernames will be overwritten.",
                     font=ctk.CTkFont(size=11), text_color="gray60").pack(
            anchor="w", padx=8, pady=(4, 0))

    # -- Export helpers ──────────────────────────────────────────────────────
    def _export_json(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".json", filetypes=[("JSON file", "*.json")])
        if not path:
            return
        with open(path, "w") as fp:
            json.dump(self.users, fp, indent=4)
        messagebox.showinfo("Exported", f"Saved {len(self.users)} users → {path}")

    def _export_csv(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV file", "*.csv")])
        if not path:
            return
        headers = ["username", "full_name", "email", "age",
                   "occupation", "location", "role", "date_added"]
        with open(path, "w", newline="", encoding="utf-8") as fp:
            w = csv.DictWriter(fp, fieldnames=headers, extrasaction="ignore")
            w.writeheader()
            for uname, d in self.users.items():
                row = {"username": uname}
                row.update({k: d.get(k, "") for k in headers[1:]})
                w.writerow(row)
        messagebox.showinfo("Exported", f"Saved {len(self.users)} users → {path}")

    def _export_excel(self):
        if not EXCEL_OK:
            messagebox.showerror("Missing Library",
                                 "Install openpyxl first:\n  pip install openpyxl")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel file", "*.xlsx")])
        if not path:
            return
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Users"

        header_fill  = PatternFill("solid", fgColor="1f3d7a")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        headers      = ["Username", "Full Name", "Email", "Age",
                        "Occupation", "Location", "Role", "Date Added"]
        keys         = ["", "full_name", "email", "age",
                        "occupation", "location", "role", "date_added"]

        for col, (h, _) in enumerate(zip(headers, keys), start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_i, (uname, d) in enumerate(self.users.items(), start=2):
            ws.cell(row=row_i, column=1, value=uname)
            for col_i, key in enumerate(keys[1:], start=2):
                ws.cell(row=row_i, column=col_i, value=d.get(key, ""))

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        wb.save(path)
        messagebox.showinfo("Exported", f"Saved {len(self.users)} users → {path}")

    # -- Import helpers ──────────────────────────────────────────────────────
    def _import_json(self):
        path = filedialog.askopenfilename(filetypes=[("JSON file", "*.json")])
        if not path:
            return
        try:
            with open(path, "r") as fp:
                raw = json.load(fp)
            count = 0
            if isinstance(raw, dict):
                self.users.update(raw)
                count = len(raw)
            elif isinstance(raw, list):
                for u in raw:
                    uname = u.get("username") or u.get("name")
                    if uname:
                        self.users[uname] = {
                            k: v for k, v in u.items()
                            if k not in ("username",)
                        }
                        count += 1
            save_users(self.users)
            messagebox.showinfo("Imported", f"Imported {count} users from {path}")
        except Exception as e:
            messagebox.showerror("Import Error", str(e))

    def _import_csv(self):
        path = filedialog.askopenfilename(filetypes=[("CSV file", "*.csv")])
        if not path:
            return
        try:
            count = 0
            with open(path, "r", encoding="utf-8") as fp:
                for row in csv.DictReader(fp):
                    uname = row.get("username") or row.get("name")
                    if uname:
                        self.users[uname] = {
                            k: v for k, v in row.items()
                            if k not in ("username", "name")
                        }
                        self.users[uname].setdefault("promotion_requests", [])
                        count += 1
            save_users(self.users)
            messagebox.showinfo("Imported", f"Imported {count} users from {path}")
        except Exception as e:
            messagebox.showerror("Import Error", str(e))


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = UserManagerPro()
    app.mainloop()
