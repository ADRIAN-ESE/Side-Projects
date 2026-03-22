-"""
╔══════════════════════════════════════════════════════════════╗
║          ARCHER ENTERPRISE — Corporate Management System      ║
║          Desktop Edition  |  CustomTkinter + Matplotlib       ║
╠══════════════════════════════════════════════════════════════╣
║  pip install customtkinter openpyxl matplotlib               ║
║  python archer_desktop.py                                     ║
╠══════════════════════════════════════════════════════════════╣
║  Shares archer_data.json with the Web Edition.               ║
║  DEFAULT ACCOUNTS:                                            ║
║    admin / admin123  |  hr_manager / hr123                    ║
║    manager1 / mgr123 |  staff1 / staff123                     ║
╚══════════════════════════════════════════════════════════════╝
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json, os, hashlib, csv, base64
from datetime import datetime

# ── Optional dependencies ─────────────────────────────────────────────────
try:
    import matplotlib
    matplotlib.use("TkAgg")
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    import matplotlib.pyplot as plt
    MPL_OK = True
except ImportError:
    MPL_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── Config ────────────────────────────────────────────────────────────────
DATA_FILE = "archer_data.json"
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

GOLD     = "#c9a84c"
GOLD_L   = "#e8c56d"
GOLD_DIM = "#1a1608"
BG       = "#07091a"
SURFACE  = "#0b0e20"
CARD     = "#101426"
BORDER   = "#1c2040"
TEXT     = "#e2e4f0"
MUTED    = "#4e5475"
SUCCESS  = "#3dba7a"
WARNING  = "#e09a30"
DANGER   = "#e05252"
INFO     = "#4f8ef7"

ROLE_LEVEL = {"Admin": 4, "HR": 3, "Manager": 2, "Staff": 1}


# ════════════════════════════════════════════════════════════
#  DATA HELPERS  (shared with web edition via archer_data.json)
# ════════════════════════════════════════════════════════════
def hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()


def init_data() -> dict:
    d = {
        "meta": {"last_emp_num": 0},
        "system_users": {
            "admin":      {"password_hash": hash_pw("admin123"), "role": "Admin",   "full_name": "System Administrator", "emp_id": None},
            "hr_manager": {"password_hash": hash_pw("hr123"),    "role": "HR",      "full_name": "HR Manager",            "emp_id": None},
            "manager1":   {"password_hash": hash_pw("mgr123"),   "role": "Manager", "full_name": "Department Manager",    "emp_id": None},
            "staff1":     {"password_hash": hash_pw("staff123"), "role": "Staff",   "full_name": "Staff Member",          "emp_id": None},
        },
        "employees": {},
        "departments": {
            "Executive":       {"description": "Executive leadership team",   "head_id": ""},
            "Human Resources": {"description": "HR and recruitment",          "head_id": ""},
            "Engineering":     {"description": "Software and systems eng.",   "head_id": ""},
            "Finance":         {"description": "Financial operations",        "head_id": ""},
            "Operations":      {"description": "Day-to-day operations",       "head_id": ""},
            "Sales":           {"description": "Sales & business development","head_id": ""},
        },
        "audit_log": []
    }
    save_data(d)
    return d


def load_data() -> dict:
    if not os.path.exists(DATA_FILE):
        return init_data()
    try:
        with open(DATA_FILE) as f:
            return json.load(f)
    except Exception:
        return init_data()


def save_data(d: dict) -> None:
    with open(DATA_FILE, "w") as f:
        json.dump(d, f, indent=4)


def next_id(d: dict) -> str:
    d["meta"]["last_emp_num"] += 1
    return f"ARCH-{datetime.now().year}-{d['meta']['last_emp_num']:04d}"


def add_audit(d: dict, user: str, action: str, target: str, detail: str) -> None:
    d["audit_log"].insert(0, {
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user": user, "action": action, "target": target, "detail": detail
    })
    d["audit_log"] = d["audit_log"][:500]


# ════════════════════════════════════════════════════════════
#  LOGIN WINDOW
# ════════════════════════════════════════════════════════════
class LoginWindow(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("ARCHER ENTERPRISE — Login")
        self.geometry("460x580")
        self.resizable(False, False)
        self.configure(fg_color=BG)
        self._build()

    def _build(self):
        # Logo section
        ctk.CTkLabel(self, text="🏹", font=ctk.CTkFont(size=40)).pack(pady=(50, 4))
        ctk.CTkLabel(self, text="ARCHER ENTERPRISE",
                     font=ctk.CTkFont(family="Helvetica", size=26, weight="bold"),
                     text_color=GOLD).pack()
        ctk.CTkLabel(self, text="CORPORATE MANAGEMENT & SERVICE AGENCY",
                     font=ctk.CTkFont(size=9),
                     text_color=MUTED).pack(pady=(2, 36))

        # Card
        card = ctk.CTkFrame(self, fg_color=CARD,
                            border_color=BORDER, border_width=1,
                            corner_radius=14, width=360)
        card.pack(padx=50)
        card.pack_propagate(False)

        ctk.CTkLabel(card, text="SYSTEM ACCESS",
                     font=ctk.CTkFont(family="Helvetica", size=16, weight="bold"),
                     text_color="white").pack(pady=(26, 2))
        ctk.CTkLabel(card, text="Authorized personnel only",
                     font=ctk.CTkFont(size=11), text_color=MUTED).pack(pady=(0, 22))

        self.err_lbl = ctk.CTkLabel(card, text="", font=ctk.CTkFont(size=12),
                                     text_color=DANGER)
        self.err_lbl.pack(pady=(0, 6))

        for label, attr, show in [("USERNAME", "username_entry", None),
                                   ("PASSWORD", "password_entry", "•")]:
            ctk.CTkLabel(card, text=label,
                         font=ctk.CTkFont(size=10, weight="bold"),
                         text_color=MUTED, anchor="w").pack(anchor="w", padx=26)
            entry = ctk.CTkEntry(card, height=40, corner_radius=8,
                                  fg_color="#0a0d1e", border_color=BORDER,
                                  text_color=TEXT,
                                  font=ctk.CTkFont(size=14),
                                  show=show or "")
            entry.pack(fill="x", padx=26, pady=(4, 14))
            setattr(self, attr, entry)

        self.password_entry.bind("<Return>", lambda e: self._login())

        ctk.CTkButton(card, text="ACCESS SYSTEM", height=44, corner_radius=8,
                      fg_color=GOLD, hover_color=GOLD_L,
                      text_color="#07091a",
                      font=ctk.CTkFont(family="Helvetica", size=14, weight="bold"),
                      command=self._login).pack(fill="x", padx=26, pady=(4, 26))

        # Default credentials helper
        creds_frame = ctk.CTkFrame(self, fg_color="transparent")
        creds_frame.pack(padx=50, pady=(16, 0), fill="x")
        ctk.CTkLabel(creds_frame, text="DEMO CREDENTIALS",
                     font=ctk.CTkFont(size=9, weight="bold"),
                     text_color=GOLD).pack(anchor="w")
        for user, pw, role in [("admin","admin123","Admin"),("hr_manager","hr123","HR"),
                                ("manager1","mgr123","Manager"),("staff1","staff123","Staff")]:
            row = ctk.CTkFrame(creds_frame, fg_color="transparent")
            row.pack(fill="x", pady=1)
            ctk.CTkLabel(row, text=f"{user} / {pw}",
                         font=ctk.CTkFont(family="Courier New", size=11),
                         text_color=TEXT).pack(side="left")
            ctk.CTkLabel(row, text=role,
                         font=ctk.CTkFont(size=10, weight="bold"),
                         text_color=GOLD).pack(side="right")

    def _login(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get()
        d = load_data()
        user = d["system_users"].get(username)
        if user and user["password_hash"] == hash_pw(password):
            add_audit(d, username, "LOGIN", username, f"Desktop login as {user['role']}")
            save_data(d)
            self.destroy()
            app = ArcherApp(username, user["role"], user.get("full_name", username))
            app.mainloop()
        else:
            self.err_lbl.configure(text="❌  Invalid username or password")


# ════════════════════════════════════════════════════════════
#  EMPLOYEE DIALOG
# ════════════════════════════════════════════════════════════
class EmpDialog(ctk.CTkToplevel):
    FIELDS = [
        ("Full Name *",       "full_name",       ""),
        ("Email *",           "email",           ""),
        ("Department *",      "department",      "__DEPT_SELECT__"),
        ("Role / Level",      "role",            "__ROLE_SELECT__"),
        ("Occupation / Title","occupation",      ""),
        ("Age",               "age",             ""),
        ("Location",          "location",        ""),
        ("Employment Type",   "employment_type", "__TYPE_SELECT__"),
        ("Status",            "status",          "__STATUS_SELECT__"),
        ("Salary",            "salary",          ""),
    ]
    ROLES  = ["Staff","Senior Staff","Manager","Senior Manager","Director","VP","C-Level"]
    TYPES  = ["Full-Time","Part-Time","Contract","Intern"]

    def __init__(self, parent, emp_id: str, data: dict, departments: list, on_save):
        super().__init__(parent)
        self.title(f"{'Edit' if emp_id else 'Add'} Employee")
        self.geometry("560x700")
        self.grab_set()
        self.configure(fg_color=BG)

        self.emp_id  = emp_id
        self.data    = dict(data)
        self.depts   = departments
        self.on_save = on_save
        self.photo_data = data.get("photo", "")
        self.vars: dict[str, tk.StringVar] = {}
        self._build()

    def _build(self):
        ctk.CTkLabel(self,
                     text=f"{'EDIT' if self.emp_id else 'ADD'} EMPLOYEE" + (f" — {self.emp_id}" if self.emp_id else ""),
                     font=ctk.CTkFont(family="Helvetica", size=17, weight="bold"),
                     text_color=GOLD).pack(pady=(22, 4), padx=24, anchor="w")

        scr = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scr.pack(fill="both", expand=True, padx=20, pady=4)
        scr.columnconfigure(1, weight=1)

        # Photo upload
        ctk.CTkLabel(scr, text="PROFILE PHOTO",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     text_color=MUTED).grid(row=0, column=0, sticky="nw", padx=8, pady=8)
        photo_btn = ctk.CTkButton(scr, text="📷 Upload Photo", width=160, height=34,
                                   fg_color=SURFACE, hover_color=CARD,
                                   border_color=BORDER, border_width=1,
                                   command=self._pick_photo)
        photo_btn.grid(row=0, column=1, padx=8, pady=8, sticky="w")
        self.photo_lbl = ctk.CTkLabel(scr, text="No photo" if not self.photo_data else "Photo loaded ✓",
                                       font=ctk.CTkFont(size=11),
                                       text_color=SUCCESS if self.photo_data else MUTED)
        self.photo_lbl.grid(row=0, column=1, padx=(180, 8), pady=8, sticky="w")

        for i, (label, key, hint) in enumerate(self.FIELDS, start=1):
            ctk.CTkLabel(scr, text=label,
                         font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=MUTED).grid(row=i, column=0, sticky="w", padx=8, pady=7)
            var = tk.StringVar(value=self.data.get(key, ""))
            self.vars[key] = var

            if hint == "__DEPT_SELECT__":
                combo = ctk.CTkComboBox(scr, variable=var, values=self.depts, height=34,
                                         fg_color="#0a0d1e", border_color=BORDER)
                combo.grid(row=i, column=1, padx=8, pady=7, sticky="ew")
            elif hint == "__ROLE_SELECT__":
                combo = ctk.CTkComboBox(scr, variable=var, values=self.ROLES, height=34,
                                         fg_color="#0a0d1e", border_color=BORDER)
                combo.set(self.data.get(key, "Staff"))
                combo.grid(row=i, column=1, padx=8, pady=7, sticky="ew")
            elif hint == "__TYPE_SELECT__":
                combo = ctk.CTkComboBox(scr, variable=var, values=self.TYPES, height=34,
                                         fg_color="#0a0d1e", border_color=BORDER)
                combo.set(self.data.get(key, "Full-Time"))
                combo.grid(row=i, column=1, padx=8, pady=7, sticky="ew")
            elif hint == "__STATUS_SELECT__":
                combo = ctk.CTkComboBox(scr, variable=var, values=["active","inactive"], height=34,
                                         fg_color="#0a0d1e", border_color=BORDER)
                combo.set(self.data.get(key, "active"))
                combo.grid(row=i, column=1, padx=8, pady=7, sticky="ew")
            else:
                entry = ctk.CTkEntry(scr, textvariable=var, height=34,
                                      fg_color="#0a0d1e", border_color=BORDER,
                                      font=ctk.CTkFont(size=13))
                entry.grid(row=i, column=1, padx=8, pady=7, sticky="ew")

        ctk.CTkButton(self, text="💾  Save Employee", height=44,
                      fg_color=GOLD, hover_color=GOLD_L,
                      text_color="#07091a",
                      font=ctk.CTkFont(family="Helvetica", size=14, weight="bold"),
                      command=self._save).pack(pady=16, padx=24, fill="x")

    def _pick_photo(self):
        path = filedialog.askopenfilename(filetypes=[("Images","*.png *.jpg *.jpeg *.gif")])
        if not path:
            return
        with open(path, "rb") as f:
            ext = path.rsplit(".", 1)[-1].lower()
            mime = {"png":"image/png","jpg":"image/jpeg","jpeg":"image/jpeg","gif":"image/gif"}.get(ext,"image/jpeg")
            self.photo_data = f"data:{mime};base64," + base64.b64encode(f.read()).decode()
        self.photo_lbl.configure(text="Photo loaded ✓", text_color=SUCCESS)

    def _save(self):
        vals = {k: v.get().strip() for k, v in self.vars.items()}
        if not vals["full_name"] or not vals["email"] or not vals["department"]:
            messagebox.showerror("Missing Fields", "Full name, email, and department are required.")
            return
        vals["photo"] = self.photo_data
        self.on_save(self.emp_id, vals)
        self.destroy()


# ════════════════════════════════════════════════════════════
#  MAIN APP WINDOW
# ════════════════════════════════════════════════════════════
class ArcherApp(ctk.CTk):
    def __init__(self, username: str, role: str, full_name: str):
        super().__init__()
        self.username  = username
        self.role      = role
        self.full_name = full_name
        self.my_level  = ROLE_LEVEL.get(role, 1)
        self.title("ARCHER ENTERPRISE — Corporate Management System")
        self.geometry("1280x820")
        self.minsize(1000, 640)
        self.configure(fg_color=BG)
        self._setup_styles()
        self._build()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("A.Treeview",
                        background=CARD, foreground=TEXT,
                        fieldbackground=CARD, rowheight=32,
                        font=("Segoe UI", 11))
        style.configure("A.Treeview.Heading",
                        background="#080b1a", foreground=MUTED,
                        font=("Segoe UI", 10, "bold"), relief="flat")
        style.map("A.Treeview",
                  background=[("selected", "#1a2a4a")],
                  foreground=[("selected", GOLD)])

    # ── Build skeleton ────────────────────────────────────────────────────
    def _build(self):
        # Sidebar
        self.sidebar = ctk.CTkFrame(self, width=220, corner_radius=0,
                                     fg_color=SURFACE,
                                     border_color=BORDER, border_width=1)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self._build_sidebar()

        # Content
        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(side="right", fill="both", expand=True)

        # Topbar
        topbar = ctk.CTkFrame(content, height=56, corner_radius=0,
                               fg_color=SURFACE,
                               border_color=BORDER, border_width=1)
        topbar.pack(side="top", fill="x")
        topbar.pack_propagate(False)
        self.page_title_lbl = ctk.CTkLabel(
            topbar, text="DASHBOARD",
            font=ctk.CTkFont(family="Helvetica", size=20, weight="bold"),
            text_color="white")
        self.page_title_lbl.pack(side="left", padx=24)
        ctk.CTkLabel(topbar,
                     text=datetime.now().strftime("%A, %d %B %Y"),
                     font=ctk.CTkFont(size=11), text_color=MUTED).pack(side="right", padx=16)
        ctk.CTkLabel(topbar, text=f"  {self.role.upper()}  ",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     text_color=GOLD, fg_color="#1a1608",
                     corner_radius=6).pack(side="right", padx=4)

        # Main area
        self.main = ctk.CTkFrame(content, fg_color="transparent")
        self.main.pack(fill="both", expand=True, padx=22, pady=18)

        # Build pages
        self.pages: dict[str, ctk.CTkFrame] = {}
        self._build_dashboard()
        self._build_employees()
        self._build_departments()
        self._build_promotions()
        self._build_audit()
        self._build_profile()

        self.show_page("dashboard")

    def _build_sidebar(self):
        # Logo
        logo_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent",
                                   border_color=BORDER, border_width=0)
        logo_frame.pack(fill="x", padx=0, pady=0)
        ctk.CTkFrame(self.sidebar, height=1, fg_color=BORDER).pack(fill="x")

        logo_inner = ctk.CTkFrame(logo_frame, fg_color="transparent")
        logo_inner.pack(padx=18, pady=(22, 16))
        ctk.CTkLabel(logo_inner, text="🏹  ARCHER",
                     font=ctk.CTkFont(family="Helvetica", size=20, weight="bold"),
                     text_color=GOLD).pack(anchor="w")
        ctk.CTkLabel(logo_inner, text="ENTERPRISE",
                     font=ctk.CTkFont(size=9, weight="bold"),
                     text_color=MUTED).pack(anchor="w")

        ctk.CTkFrame(self.sidebar, height=1, fg_color=BORDER).pack(fill="x")

        # Nav
        nav_items = [
            ("📊", "Dashboard",     "dashboard",   1),
            ("👥", "Employees",     "employees",   1),
            ("🏢", "Departments",   "departments", 2),
            ("🏆", "Promotions",    "promotions",  2),
            ("📋", "Audit Log",     "audit",       3),
            ("🧑", "My Profile",    "profile",     1),
        ]
        ctk.CTkLabel(self.sidebar, text="WORKSPACE",
                     font=ctk.CTkFont(size=9, weight="bold"),
                     text_color=MUTED).pack(anchor="w", padx=18, pady=(14, 4))

        self.nav_buttons: dict[str, ctk.CTkButton] = {}
        for icon, label, key, min_level in nav_items:
            if self.my_level < min_level:
                continue
            btn = ctk.CTkButton(
                self.sidebar,
                text=f"  {icon}  {label}",
                anchor="w", height=40, width=210,
                fg_color="transparent",
                hover_color="#141728",
                text_color=MUTED,
                font=ctk.CTkFont(size=13),
                command=lambda k=key: self.show_page(k),
            )
            btn.pack(pady=1, padx=8)
            self.nav_buttons[key] = btn

        # User section
        ctk.CTkFrame(self.sidebar, height=1, fg_color=BORDER).pack(fill="x", side="bottom", pady=(0, 0))
        user_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        user_frame.pack(side="bottom", fill="x", padx=16, pady=16)

        initials = "".join(w[0] for w in self.full_name.split()[:2]).upper() if self.full_name else "?"
        ctk.CTkLabel(user_frame, text=initials, width=36, height=36,
                     fg_color=GOLD, text_color="#07091a", corner_radius=18,
                     font=ctk.CTkFont(family="Helvetica", size=14, weight="bold")).pack(side="left")

        info = ctk.CTkFrame(user_frame, fg_color="transparent")
        info.pack(side="left", padx=10, fill="x", expand=True)
        ctk.CTkLabel(info, text=self.full_name,
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color="white", anchor="w").pack(anchor="w")
        ctk.CTkLabel(info, text=self.role,
                     font=ctk.CTkFont(size=10),
                     text_color=GOLD, anchor="w").pack(anchor="w")

        ctk.CTkButton(self.sidebar, text="← Sign Out", height=34, width=210,
                      fg_color="transparent",
                      hover_color="#2a0a0a",
                      text_color=DANGER,
                      font=ctk.CTkFont(size=12),
                      command=self._logout).pack(side="bottom", padx=8, pady=(0, 8))

    # ── Page routing ──────────────────────────────────────────────────────
    def show_page(self, name: str):
        for page in self.pages.values():
            page.pack_forget()
        if name in self.pages:
            self.pages[name].pack(fill="both", expand=True)
        for key, btn in self.nav_buttons.items():
            is_active = key == name
            btn.configure(
                fg_color="#1a1d30" if is_active else "transparent",
                text_color=GOLD if is_active else MUTED,
                border_width=0,
            )
        titles = {"dashboard":"DASHBOARD","employees":"EMPLOYEES","departments":"DEPARTMENTS",
                  "promotions":"PROMOTIONS","audit":"AUDIT LOG","profile":"MY PROFILE"}
        self.page_title_lbl.configure(text=titles.get(name, name.upper()))
        refreshers = {
            "dashboard":   self._refresh_dashboard,
            "employees":   self._refresh_emp_tree,
            "departments": self._refresh_depts,
            "promotions":  self._refresh_promos,
            "audit":       self._refresh_audit,
            "profile":     self._refresh_profile,
        }
        if name in refreshers:
            refreshers[name]()

    # ══════════════════════════════════════════════════════
    #  DASHBOARD
    # ══════════════════════════════════════════════════════
    def _build_dashboard(self):
        f = ctk.CTkFrame(self.main, fg_color="transparent")
        self.pages["dashboard"] = f

        # Stat cards row
        self.stat_row = ctk.CTkFrame(f, fg_color="transparent")
        self.stat_row.pack(fill="x", pady=(0, 16))

        # Charts row (matplotlib)
        self.chart_row = ctk.CTkFrame(f, fg_color="transparent")
        self.chart_row.pack(fill="both", expand=True)

    def _refresh_dashboard(self):
        d = load_data()
        emps = d["employees"]
        depts = d["departments"]
        active = sum(1 for e in emps.values() if e.get("status") == "active")
        pending = sum(
            1 for e in emps.values()
            for r in e.get("promotion_requests", [])
            if r.get("status") == "pending"
        )

        for w in self.stat_row.winfo_children():
            w.destroy()

        for label, value, color in [
            ("TOTAL EMPLOYEES", len(emps), GOLD),
            ("ACTIVE STAFF",    active,    SUCCESS),
            ("PENDING PROMOS",  pending,   WARNING),
            ("DEPARTMENTS",     len(depts),INFO),
        ]:
            card = ctk.CTkFrame(self.stat_row, fg_color=CARD,
                                 border_color=BORDER, border_width=1,
                                 corner_radius=10)
            card.pack(side="left", fill="both", expand=True, padx=5)
            ctk.CTkFrame(card, height=3, fg_color=color, corner_radius=2).pack(fill="x")
            ctk.CTkLabel(card, text=str(value),
                         font=ctk.CTkFont(family="Helvetica", size=40, weight="bold"),
                         text_color="white").pack(pady=(14, 2))
            ctk.CTkLabel(card, text=label,
                         font=ctk.CTkFont(size=9, weight="bold"),
                         text_color=MUTED).pack(pady=(0, 14))

        # Matplotlib charts
        for w in self.chart_row.winfo_children():
            w.destroy()

        if not MPL_OK:
            ctk.CTkLabel(self.chart_row,
                         text="Install matplotlib for charts:\n  pip install matplotlib",
                         text_color=MUTED, font=ctk.CTkFont(size=13)).pack(expand=True)
            return

        dept_counts = {}
        role_counts = {}
        for e in emps.values():
            dept_counts[e.get("department","?")] = dept_counts.get(e.get("department","?"),0)+1
            role_counts[e.get("role","Staff")]   = role_counts.get(e.get("role","Staff"),0)+1

        GOLD_PAL = ["#c9a84c","#e8c56d","#a07a30","#f0d88a","#8a6020","#d4aa60","#b89040"]

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 3.2),
                                        facecolor=CARD)
        for ax in (ax1, ax2):
            ax.set_facecolor(CARD)
            for spine in ax.spines.values():
                spine.set_edgecolor(BORDER)

        # Bar chart — employees by dept
        if dept_counts:
            bars = ax1.bar(dept_counts.keys(), dept_counts.values(),
                           color=[GOLD_PAL[i % len(GOLD_PAL)] for i in range(len(dept_counts))],
                           edgecolor="none", width=0.6)
            ax1.set_title("STAFF BY DEPARTMENT", color=TEXT, fontsize=9, fontweight="bold", pad=10)
            ax1.tick_params(colors=MUTED, labelsize=8)
            ax1.set_xlabel("", color=MUTED)
            plt.setp(ax1.get_xticklabels(), rotation=30, ha="right", fontsize=7)
            for bar in bars:
                ax1.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.05,
                         str(int(bar.get_height())), ha="center", va="bottom",
                         fontsize=8, color=GOLD)
        else:
            ax1.text(0.5, 0.5, "No data", ha="center", va="center",
                     color=MUTED, transform=ax1.transAxes)
            ax1.set_title("STAFF BY DEPARTMENT", color=TEXT, fontsize=9, fontweight="bold")

        # Pie — role distribution
        if role_counts:
            wedges, _, autotexts = ax2.pie(
                role_counts.values(), labels=role_counts.keys(),
                colors=GOLD_PAL[:len(role_counts)],
                autopct="%1.0f%%", startangle=90,
                wedgeprops={"linewidth":2, "edgecolor":CARD},
                pctdistance=0.8
            )
            for t in autotexts: t.set_color(CARD); t.set_fontsize(8)
            for t in ax2.texts: t.set_color(MUTED); t.set_fontsize(8)
        ax2.set_title("ROLE DISTRIBUTION", color=TEXT, fontsize=9, fontweight="bold", pad=10)

        fig.tight_layout(pad=1.5)
        canvas = FigureCanvasTkAgg(fig, master=self.chart_row)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig)

    # ══════════════════════════════════════════════════════
    #  EMPLOYEES
    # ══════════════════════════════════════════════════════
    def _build_employees(self):
        f = ctk.CTkFrame(self.main, fg_color="transparent")
        self.pages["employees"] = f

        # Toolbar
        toolbar = ctk.CTkFrame(f, fg_color="transparent")
        toolbar.pack(fill="x", pady=(0, 10))

        self.emp_search_var = tk.StringVar()
        self.emp_search_var.trace_add("write", lambda *_: self._refresh_emp_tree())
        ctk.CTkEntry(toolbar, textvariable=self.emp_search_var,
                      placeholder_text="🔍  Search name, ID, department, role…",
                      width=360, height=36,
                      fg_color=CARD, border_color=BORDER,
                      font=ctk.CTkFont(size=13)).pack(side="left")

        if self.my_level >= 3:
            ctk.CTkButton(toolbar, text="＋ Add Employee", height=36,
                          fg_color=GOLD, hover_color=GOLD_L,
                          text_color="#07091a",
                          font=ctk.CTkFont(size=13, weight="bold"),
                          command=self._open_add_emp).pack(side="right", padx=4)
        ctk.CTkButton(toolbar, text="⬇ CSV", height=36, width=80,
                      fg_color=SURFACE, hover_color=CARD,
                      border_color=BORDER, border_width=1,
                      font=ctk.CTkFont(size=12),
                      command=self._export_csv).pack(side="right", padx=4)
        if EXCEL_OK:
            ctk.CTkButton(toolbar, text="⬇ Excel", height=36, width=90,
                          fg_color=SURFACE, hover_color=CARD,
                          border_color=BORDER, border_width=1,
                          font=ctk.CTkFont(size=12),
                          command=self._export_excel).pack(side="right", padx=4)

        # Treeview
        cols = ("ID","Full Name","Department","Role","Occupation","Email","Status","Joined")
        self.emp_tree = ttk.Treeview(f, columns=cols, show="headings",
                                      selectmode="browse", style="A.Treeview")
        col_widths = {"ID":140,"Full Name":180,"Department":140,"Role":120,
                      "Occupation":140,"Email":190,"Status":80,"Joined":100}
        for col in cols:
            self.emp_tree.heading(col, text=col)
            self.emp_tree.column(col, width=col_widths.get(col,120), minwidth=60)

        vsb = ttk.Scrollbar(f, orient="vertical", command=self.emp_tree.yview)
        hsb = ttk.Scrollbar(f, orient="horizontal", command=self.emp_tree.xview)
        self.emp_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.emp_tree.grid(row=1, column=0, sticky="nsew", in_=f)
        vsb.grid(row=1, column=1, sticky="ns", in_=f)
        hsb.grid(row=2, column=0, sticky="ew", in_=f)
        f.rowconfigure(1, weight=1); f.columnconfigure(0, weight=1)

        # Bottom action bar
        act_bar = ctk.CTkFrame(f, fg_color="transparent")
        act_bar.grid(row=3, column=0, sticky="ew", pady=(10,0))
        if self.my_level >= 3:
            ctk.CTkButton(act_bar, text="✏️  Edit Selected", height=36, width=160,
                          fg_color=SURFACE, hover_color=CARD,
                          border_color=BORDER, border_width=1,
                          command=self._open_edit_emp).pack(side="left", padx=4)
        if self.my_level >= 4:
            ctk.CTkButton(act_bar, text="🗑  Delete Selected", height=36, width=160,
                          fg_color="transparent",
                          hover_color="#2a0a0a",
                          border_color=DANGER, border_width=1,
                          text_color=DANGER,
                          command=self._delete_emp).pack(side="left", padx=4)

    def _refresh_emp_tree(self):
        q = getattr(self, "emp_search_var", None)
        q = q.get().lower().strip() if q else ""
        d = load_data()
        emps = d["employees"]

        # Staff see only their own record
        if self.my_level == 1:
            su = d["system_users"].get(self.username, {})
            emp_id = su.get("emp_id")
            emps = {emp_id: emps[emp_id]} if emp_id and emp_id in emps else {}

        for row in self.emp_tree.get_children():
            self.emp_tree.delete(row)

        for emp_id, emp in emps.items():
            if q and not any(q in str(v).lower() for v in [
                emp_id, emp.get("full_name",""), emp.get("department",""),
                emp.get("role",""), emp.get("email",""), emp.get("occupation","")
            ]):
                continue
            self.emp_tree.insert("", "end", iid=emp_id, values=(
                emp_id, emp.get("full_name",""), emp.get("department",""),
                emp.get("role",""), emp.get("occupation",""), emp.get("email",""),
                emp.get("status","active"), emp.get("date_added",""),
            ))

    def _open_add_emp(self):
        d = load_data()
        EmpDialog(self, "", {}, list(d["departments"].keys()), self._save_emp)

    def _open_edit_emp(self):
        sel = self.emp_tree.selection()
        if not sel: messagebox.showwarning("No Selection","Select an employee first."); return
        emp_id = sel[0]
        d = load_data()
        EmpDialog(self, emp_id, d["employees"][emp_id], list(d["departments"].keys()), self._save_emp)

    def _save_emp(self, emp_id: str, vals: dict):
        d = load_data()
        if emp_id:
            d["employees"][emp_id].update(vals)
            add_audit(d, self.username, "EDIT_EMPLOYEE", emp_id, f"Updated {vals['full_name']}")
        else:
            new_id = next_id(d)
            vals["promotion_requests"] = []
            vals["date_added"] = datetime.now().strftime("%Y-%m-%d")
            vals.setdefault("status","active")
            d["employees"][new_id] = vals
            add_audit(d, self.username, "ADD_EMPLOYEE", new_id, f"Added {vals['full_name']} to {vals.get('department','')}")
        save_data(d)
        self._refresh_emp_tree()
        messagebox.showinfo("Saved", f"Employee record saved.")

    def _delete_emp(self):
        sel = self.emp_tree.selection()
        if not sel: messagebox.showwarning("No Selection","Select an employee first."); return
        emp_id = sel[0]
        if not messagebox.askyesno("Confirm Delete", f"Delete employee {emp_id}?\nThis cannot be undone."): return
        d = load_data()
        name = d["employees"].get(emp_id,{}).get("full_name","")
        del d["employees"][emp_id]
        add_audit(d, self.username, "DELETE_EMPLOYEE", emp_id, f"Removed {name}")
        save_data(d)
        self._refresh_emp_tree()

    # ══════════════════════════════════════════════════════
    #  DEPARTMENTS
    # ══════════════════════════════════════════════════════
    def _build_departments(self):
        f = ctk.CTkScrollableFrame(self.main, fg_color="transparent")
        self.pages["departments"] = f

        toolbar = ctk.CTkFrame(f, fg_color="transparent")
        toolbar.pack(fill="x", pady=(0,12))
        ctk.CTkLabel(toolbar, text="DEPARTMENTS",
                     font=ctk.CTkFont(family="Helvetica",size=16,weight="bold"),
                     text_color="white").pack(side="left")
        if self.my_level >= 3:
            ctk.CTkButton(toolbar, text="＋ Add Department", height=36,
                          fg_color=GOLD, hover_color=GOLD_L,
                          text_color="#07091a",
                          font=ctk.CTkFont(size=12,weight="bold"),
                          command=self._add_dept_dialog).pack(side="right")

        self.dept_content = ctk.CTkFrame(f, fg_color="transparent")
        self.dept_content.pack(fill="both", expand=True)

    def _refresh_depts(self):
        for w in self.dept_content.winfo_children():
            w.destroy()
        d = load_data()
        for i, (name, info) in enumerate(d["departments"].items()):
            count = sum(1 for e in d["employees"].values() if e.get("department")==name)
            head  = info.get("head_id","")
            head_name = d["employees"].get(head,{}).get("full_name",head) if head else "Unassigned"

            card = ctk.CTkFrame(self.dept_content, fg_color=CARD,
                                 border_color=BORDER, border_width=1, corner_radius=10)
            card.pack(fill="x", pady=5)

            top_row = ctk.CTkFrame(card, fg_color="transparent")
            top_row.pack(fill="x", padx=20, pady=(16,4))
            ctk.CTkLabel(top_row, text=name,
                         font=ctk.CTkFont(family="Helvetica",size=16,weight="bold"),
                         text_color="white").pack(side="left")
            ctk.CTkLabel(top_row, text=f"{count} employees",
                         font=ctk.CTkFont(family="Courier New",size=14),
                         text_color=GOLD).pack(side="right")

            ctk.CTkLabel(card, text=info.get("description",""),
                         font=ctk.CTkFont(size=12), text_color=MUTED, anchor="w").pack(
                fill="x", padx=20, pady=(0,4))
            ctk.CTkLabel(card, text=f"Head: {head_name}",
                         font=ctk.CTkFont(size=11), text_color=MUTED, anchor="w").pack(
                fill="x", padx=20, pady=(0,12))

            if self.my_level >= 3:
                btn_row = ctk.CTkFrame(card, fg_color="transparent")
                btn_row.pack(fill="x", padx=20, pady=(0,12))
                ctk.CTkButton(btn_row, text="✏️ Edit", height=30, width=90,
                              fg_color=SURFACE, hover_color=CARD,
                              border_color=BORDER, border_width=1,
                              command=lambda n=name: self._edit_dept_dialog(n)).pack(side="left", padx=(0,6))
                ctk.CTkButton(btn_row, text="🗑 Delete", height=30, width=90,
                              fg_color="transparent", hover_color="#2a0a0a",
                              border_color=DANGER, border_width=1,
                              text_color=DANGER,
                              command=lambda n=name: self._delete_dept(n)).pack(side="left")

    def _add_dept_dialog(self):
        self._dept_dialog("", {})

    def _edit_dept_dialog(self, name):
        d = load_data()
        self._dept_dialog(name, d["departments"].get(name, {}))

    def _dept_dialog(self, existing_name, data):
        win = ctk.CTkToplevel(self)
        win.title("Department")
        win.geometry("420x320")
        win.grab_set()
        win.configure(fg_color=BG)

        ctk.CTkLabel(win, text="ADD DEPARTMENT" if not existing_name else f"EDIT: {existing_name}",
                     font=ctk.CTkFont(family="Helvetica",size=16,weight="bold"),
                     text_color=GOLD).pack(pady=(22,16), padx=24, anchor="w")

        vars_map = {}
        for label, key, default in [("Department Name *","name",existing_name),
                                      ("Description","description",data.get("description","")),
                                      ("Head (Employee ID)","head_id",data.get("head_id",""))]:
            ctk.CTkLabel(win, text=label,
                         font=ctk.CTkFont(size=10,weight="bold"),
                         text_color=MUTED).pack(anchor="w", padx=24)
            v = tk.StringVar(value=default)
            e = ctk.CTkEntry(win, textvariable=v, height=36,
                              fg_color="#0a0d1e", border_color=BORDER,
                              state="disabled" if existing_name and key=="name" else "normal")
            e.pack(fill="x", padx=24, pady=(4,12))
            vars_map[key] = v

        def save():
            name = vars_map["name"].get().strip()
            if not name: messagebox.showerror("Error","Name required."); return
            d = load_data()
            if not existing_name and name in d["departments"]:
                messagebox.showerror("Error","Department already exists."); return
            key = existing_name or name
            d["departments"][key] = {
                "description": vars_map["description"].get().strip(),
                "head_id": vars_map["head_id"].get().strip()
            }
            if not existing_name:
                d["departments"][name] = d["departments"].pop(key)
            add_audit(d, self.username,
                      "EDIT_DEPT" if existing_name else "ADD_DEPT",
                      name, f"{'Updated' if existing_name else 'Created'} department '{name}'")
            save_data(d)
            self._refresh_depts()
            win.destroy()
            messagebox.showinfo("Saved", f"Department '{name}' saved.")

        ctk.CTkButton(win, text="💾 Save", height=40,
                      fg_color=GOLD, hover_color=GOLD_L,
                      text_color="#07091a",
                      font=ctk.CTkFont(size=13,weight="bold"),
                      command=save).pack(fill="x", padx=24, pady=8)

    def _delete_dept(self, name):
        if not messagebox.askyesno("Confirm", f"Delete department '{name}'?"): return
        d = load_data()
        del d["departments"][name]
        add_audit(d, self.username, "DELETE_DEPT", name, f"Removed department '{name}'")
        save_data(d)
        self._refresh_depts()

    # ══════════════════════════════════════════════════════
    #  PROMOTIONS
    # ══════════════════════════════════════════════════════
    def _build_promotions(self):
        f = ctk.CTkFrame(self.main, fg_color="transparent")
        self.pages["promotions"] = f

        if self.my_level >= 2:
            req_card = ctk.CTkFrame(f, fg_color=CARD,
                                     border_color=BORDER, border_width=1, corner_radius=10)
            req_card.pack(fill="x", pady=(0,14))
            ctk.CTkLabel(req_card, text="SUBMIT REQUEST",
                         font=ctk.CTkFont(family="Helvetica",size=14,weight="bold"),
                         text_color="white").pack(anchor="w", padx=20, pady=(16,10))

            row = ctk.CTkFrame(req_card, fg_color="transparent")
            row.pack(fill="x", padx=20, pady=(0,16))
            self.p_id   = ctk.CTkEntry(row, placeholder_text="Employee ID", width=180, height=36,
                                        fg_color="#0a0d1e", border_color=BORDER)
            self.p_id.pack(side="left", padx=(0,8))
            self.p_role = ctk.CTkEntry(row, placeholder_text="Requested Level", width=180, height=36,
                                        fg_color="#0a0d1e", border_color=BORDER)
            self.p_role.pack(side="left", padx=(0,8))
            self.p_note = ctk.CTkEntry(row, placeholder_text="Justification…", width=220, height=36,
                                        fg_color="#0a0d1e", border_color=BORDER)
            self.p_note.pack(side="left", padx=(0,8))
            ctk.CTkButton(row, text="Submit", height=36, width=100,
                          fg_color=GOLD, hover_color=GOLD_L,
                          text_color="#07091a",
                          command=self._submit_promo).pack(side="left")

        cols = ("Employee ID","Name","Current Level","Requested Level","Note","Submitted By","Date","Status")
        self.promo_tree = ttk.Treeview(f, columns=cols, show="headings",
                                        selectmode="browse", style="A.Treeview")
        col_widths = {"Employee ID":130,"Name":160,"Current Level":120,"Requested Level":130,
                      "Note":180,"Submitted By":120,"Date":100,"Status":90}
        for col in cols:
            self.promo_tree.heading(col, text=col)
            self.promo_tree.column(col, width=col_widths.get(col,120), minwidth=60)
        vsb = ttk.Scrollbar(f, orient="vertical", command=self.promo_tree.yview)
        self.promo_tree.configure(yscrollcommand=vsb.set)
        self.promo_tree.pack(fill="both", expand=True, side="left")
        vsb.pack(fill="y", side="right")

        if self.my_level >= 3:
            act = ctk.CTkFrame(f, fg_color="transparent")
            act.pack(fill="x", pady=(10,0))
            ctk.CTkButton(act, text="✅ Approve Selected", height=36, width=180,
                          fg_color="transparent",
                          hover_color="#0a2a0a",
                          border_color=SUCCESS, border_width=1,
                          text_color=SUCCESS,
                          command=lambda: self._resolve_promo("approved")).pack(side="left", padx=(0,8))
            ctk.CTkButton(act, text="❌ Deny Selected", height=36, width=180,
                          fg_color="transparent",
                          hover_color="#2a0a0a",
                          border_color=DANGER, border_width=1,
                          text_color=DANGER,
                          command=lambda: self._resolve_promo("denied")).pack(side="left")

    def _refresh_promos(self):
        for row in self.promo_tree.get_children():
            self.promo_tree.delete(row)
        d = load_data()
        for emp_id, emp in d["employees"].items():
            for i, req in enumerate(emp.get("promotion_requests", [])):
                iid = f"{emp_id}::{i}"
                self.promo_tree.insert("", "end", iid=iid, values=(
                    emp_id, emp.get("full_name",""), emp.get("role",""),
                    req.get("requested_role",""), req.get("note",""),
                    req.get("submitted_by",""), req.get("date",""), req.get("status","").capitalize()
                ))

    def _submit_promo(self):
        emp_id = self.p_id.get().strip()
        role   = self.p_role.get().strip()
        note   = self.p_note.get().strip()
        d = load_data()
        if emp_id not in d["employees"]: messagebox.showerror("Error","Employee not found."); return
        if not role: messagebox.showerror("Error","Requested level is required."); return
        d["employees"][emp_id].setdefault("promotion_requests",[]).append({
            "requested_role": role, "note": note,
            "submitted_by": self.username, "status": "pending",
            "date": datetime.now().strftime("%Y-%m-%d"),
            "resolved_date":"", "resolved_by":""
        })
        add_audit(d, self.username, "PROMO_REQUEST", emp_id,
                  f"Requested promotion to '{role}' for {d['employees'][emp_id]['full_name']}")
        save_data(d)
        for entry in (self.p_id, self.p_role, self.p_note): entry.delete(0,"end")
        self._refresh_promos()

    def _resolve_promo(self, resolution):
        sel = self.promo_tree.selection()
        if not sel: messagebox.showwarning("No Selection","Select a promotion request first."); return
        emp_id, idx_s = sel[0].split("::")
        d = load_data()
        req = d["employees"][emp_id]["promotion_requests"][int(idx_s)]
        if req["status"] != "pending": messagebox.showwarning("Already Resolved","This request has already been resolved."); return
        req["status"] = resolution
        req["resolved_date"] = datetime.now().strftime("%Y-%m-%d")
        req["resolved_by"] = self.username
        if resolution == "approved":
            d["employees"][emp_id]["role"] = req["requested_role"]
        add_audit(d, self.username, f"PROMO_{resolution.upper()}", emp_id,
                  f"Promotion to '{req['requested_role']}' {resolution} for {d['employees'][emp_id]['full_name']}")
        save_data(d)
        self._refresh_promos()

    # ══════════════════════════════════════════════════════
    #  AUDIT LOG
    # ══════════════════════════════════════════════════════
    def _build_audit(self):
        f = ctk.CTkFrame(self.main, fg_color="transparent")
        self.pages["audit"] = f

        toolbar = ctk.CTkFrame(f, fg_color="transparent")
        toolbar.pack(fill="x", pady=(0,10))
        self.audit_search_var = tk.StringVar()
        self.audit_search_var.trace_add("write", lambda *_: self._refresh_audit())
        ctk.CTkEntry(toolbar, textvariable=self.audit_search_var,
                      placeholder_text="🔍  Filter log entries…",
                      width=340, height=36,
                      fg_color=CARD, border_color=BORDER).pack(side="left")

        cols = ("Timestamp","User","Action","Target","Detail")
        self.audit_tree = ttk.Treeview(f, columns=cols, show="headings",
                                        selectmode="none", style="A.Treeview")
        widths = {"Timestamp":150,"User":110,"Action":130,"Target":140,"Detail":400}
        for col in cols:
            self.audit_tree.heading(col, text=col)
            self.audit_tree.column(col, width=widths.get(col,120), minwidth=60)
        vsb = ttk.Scrollbar(f, orient="vertical", command=self.audit_tree.yview)
        self.audit_tree.configure(yscrollcommand=vsb.set)
        self.audit_tree.pack(fill="both", expand=True, side="left")
        vsb.pack(fill="y", side="right")

    def _refresh_audit(self):
        q = getattr(self,"audit_search_var",None)
        q = q.get().lower().strip() if q else ""
        for row in self.audit_tree.get_children():
            self.audit_tree.delete(row)
        d = load_data()
        for entry in d["audit_log"]:
            if q and not any(q in str(v).lower() for v in entry.values()):
                continue
            self.audit_tree.insert("", "end", values=(
                entry.get("ts",""), entry.get("user",""),
                entry.get("action",""), entry.get("target",""), entry.get("detail","")
            ))

    # ══════════════════════════════════════════════════════
    #  PROFILE
    # ══════════════════════════════════════════════════════
    def _build_profile(self):
        f = ctk.CTkScrollableFrame(self.main, fg_color="transparent")
        self.pages["profile"] = f

        # Header card
        hdr = ctk.CTkFrame(f, fg_color=CARD,
                            border_color=BORDER, border_width=1, corner_radius=10)
        hdr.pack(fill="x", pady=(0,14))

        hdr_inner = ctk.CTkFrame(hdr, fg_color="transparent")
        hdr_inner.pack(padx=24, pady=20, fill="x")

        initials = "".join(w[0] for w in self.full_name.split()[:2]).upper() if self.full_name else "?"
        ctk.CTkLabel(hdr_inner, text=initials, width=72, height=72,
                     fg_color=GOLD, text_color="#07091a", corner_radius=36,
                     font=ctk.CTkFont(family="Helvetica",size=28,weight="bold")).pack(side="left")

        info = ctk.CTkFrame(hdr_inner, fg_color="transparent")
        info.pack(side="left", padx=20)
        self.prof_name_lbl = ctk.CTkLabel(info, text=self.full_name,
                                           font=ctk.CTkFont(family="Helvetica",size=24,weight="bold"),
                                           text_color="white", anchor="w")
        self.prof_name_lbl.pack(anchor="w")
        d_now = load_data()
        emp_id = d_now["system_users"].get(self.username,{}).get("emp_id","")
        ctk.CTkLabel(info, text=emp_id or "System Account",
                     font=ctk.CTkFont(family="Courier New",size=13),
                     text_color=GOLD, anchor="w").pack(anchor="w", pady=(2,0))
        ctk.CTkLabel(info, text=self.role,
                     font=ctk.CTkFont(size=12,weight="bold"),
                     text_color=MUTED, anchor="w").pack(anchor="w")

        # Edit form
        edit_card = ctk.CTkFrame(f, fg_color=CARD,
                                  border_color=BORDER, border_width=1, corner_radius=10)
        edit_card.pack(fill="x", pady=(0,14))
        ctk.CTkLabel(edit_card, text="EDIT PROFILE",
                     font=ctk.CTkFont(family="Helvetica",size=14,weight="bold"),
                     text_color="white").pack(anchor="w", padx=24, pady=(18,12))

        form_grid = ctk.CTkFrame(edit_card, fg_color="transparent")
        form_grid.pack(fill="x", padx=24, pady=(0,16))
        form_grid.columnconfigure((0,1), weight=1)

        self.prof_vars: dict[str, tk.StringVar] = {}
        fields = [("Full Name","full_name"),("Email","email"),("Age","age"),("Location","location"),("Occupation","occupation")]
        d_data = load_data()
        su = d_data["system_users"].get(self.username, {})
        for i, (label, key) in enumerate(fields):
            row, col = divmod(i, 2)
            ctk.CTkLabel(form_grid, text=label,
                         font=ctk.CTkFont(size=10,weight="bold"),
                         text_color=MUTED).grid(row=row*2, column=col, sticky="w", padx=8, pady=(8,2))
            v = tk.StringVar(value=su.get(key,""))
            self.prof_vars[key] = v
            ctk.CTkEntry(form_grid, textvariable=v, height=36,
                          fg_color="#0a0d1e", border_color=BORDER).grid(
                row=row*2+1, column=col, sticky="ew", padx=8, pady=(0,4))

        btn_row = ctk.CTkFrame(edit_card, fg_color="transparent")
        btn_row.pack(fill="x", padx=24, pady=(0,16))
        ctk.CTkButton(btn_row, text="💾 Save Changes", height=38,
                      fg_color=GOLD, hover_color=GOLD_L,
                      text_color="#07091a",
                      font=ctk.CTkFont(size=13,weight="bold"),
                      command=self._save_profile).pack(side="left", padx=(0,8))
        ctk.CTkButton(btn_row, text="🖼 Upload Photo", height=38,
                      fg_color=SURFACE, hover_color=CARD,
                      border_color=BORDER, border_width=1,
                      command=self._upload_profile_photo).pack(side="left")

        # Change password
        pw_card = ctk.CTkFrame(f, fg_color=CARD,
                                border_color=BORDER, border_width=1, corner_radius=10)
        pw_card.pack(fill="x")
        ctk.CTkLabel(pw_card, text="CHANGE PASSWORD",
                     font=ctk.CTkFont(family="Helvetica",size=14,weight="bold"),
                     text_color="white").pack(anchor="w", padx=24, pady=(18,12))

        pw_row = ctk.CTkFrame(pw_card, fg_color="transparent")
        pw_row.pack(fill="x", padx=24, pady=(0,16))
        self.pw_vars = {}
        for label, key in [("Current Password","old"),("New Password","new"),("Confirm New","confirm")]:
            ctk.CTkLabel(pw_row, text=label,
                         font=ctk.CTkFont(size=10,weight="bold"),
                         text_color=MUTED).pack(anchor="w")
            v = tk.StringVar()
            ctk.CTkEntry(pw_row, textvariable=v, height=36, show="•",
                          fg_color="#0a0d1e", border_color=BORDER).pack(fill="x", pady=(4,10))
            self.pw_vars[key] = v

        ctk.CTkButton(pw_card, text="🔑 Change Password", height=38,
                      fg_color=GOLD, hover_color=GOLD_L,
                      text_color="#07091a",
                      font=ctk.CTkFont(size=13,weight="bold"),
                      command=self._change_password).pack(anchor="w", padx=24, pady=(0,16))

    def _refresh_profile(self): pass  # Static, no refresh needed

    def _save_profile(self):
        d = load_data()
        su = d["system_users"][self.username]
        for key, var in self.prof_vars.items():
            su[key] = var.get().strip()
        if su.get("emp_id") and su["emp_id"] in d["employees"]:
            for key, var in self.prof_vars.items():
                d["employees"][su["emp_id"]][key] = var.get().strip()
        self.full_name = su.get("full_name", self.full_name)
        self.prof_name_lbl.configure(text=self.full_name)
        add_audit(d, self.username, "PROFILE_UPDATE", self.username, "Updated own profile")
        save_data(d)
        messagebox.showinfo("Saved", "Profile updated.")

    def _upload_profile_photo(self):
        path = filedialog.askopenfilename(filetypes=[("Images","*.png *.jpg *.jpeg")])
        if not path: return
        with open(path,"rb") as f_:
            ext = path.rsplit(".",1)[-1].lower()
            mime = {"png":"image/png","jpg":"image/jpeg","jpeg":"image/jpeg"}.get(ext,"image/jpeg")
            photo = f"data:{mime};base64," + base64.b64encode(f_.read()).decode()
        d = load_data()
        d["system_users"][self.username]["photo"] = photo
        if d["system_users"][self.username].get("emp_id"):
            emp_id = d["system_users"][self.username]["emp_id"]
            if emp_id in d["employees"]:
                d["employees"][emp_id]["photo"] = photo
        add_audit(d, self.username, "PROFILE_UPDATE", self.username, "Updated profile photo")
        save_data(d)
        messagebox.showinfo("Done", "Profile photo updated.")

    def _change_password(self):
        d = load_data()
        su = d["system_users"][self.username]
        old = self.pw_vars["old"].get()
        new = self.pw_vars["new"].get()
        con = self.pw_vars["confirm"].get()
        if su["password_hash"] != hash_pw(old):
            messagebox.showerror("Error","Current password is incorrect."); return
        if new != con:
            messagebox.showerror("Error","New passwords do not match."); return
        if len(new) < 6:
            messagebox.showerror("Error","Password must be at least 6 characters."); return
        su["password_hash"] = hash_pw(new)
        add_audit(d, self.username, "PASSWORD_CHANGE", self.username, "Changed account password")
        save_data(d)
        for v in self.pw_vars.values(): v.set("")
        messagebox.showinfo("Done","Password changed successfully.")

    # ── Export ────────────────────────────────────────────────────────────
    def _export_csv(self):
        path = filedialog.asksaveasfilename(defaultextension=".csv",
                                             filetypes=[("CSV","*.csv")])
        if not path: return
        d = load_data()
        headers = ["emp_id","full_name","email","department","role","occupation",
                   "age","location","employment_type","status","salary","date_added"]
        with open(path,"w",newline="",encoding="utf-8") as fp:
            w = csv.DictWriter(fp, fieldnames=headers, extrasaction="ignore")
            w.writeheader()
            for emp_id, emp in d["employees"].items():
                row = {"emp_id": emp_id}
                row.update({k: emp.get(k,"") for k in headers[1:]})
                w.writerow(row)
        messagebox.showinfo("Exported", f"Saved {len(d['employees'])} records to:\n{path}")

    def _export_excel(self):
        if not EXCEL_OK:
            messagebox.showerror("Missing Library","Run:  pip install openpyxl"); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel","*.xlsx")])
        if not path: return
        d = load_data()
        wb = openpyxl.Workbook(); ws = wb.active; ws.title="Employees"
        hdr_fill = PatternFill("solid",fgColor="07091a")
        hdr_font = Font(bold=True,color="C9A84C",size=11)
        headers = ["Emp ID","Full Name","Email","Department","Role","Occupation",
                   "Age","Location","Type","Status","Salary","Date Added"]
        keys    = ["","full_name","email","department","role","occupation",
                   "age","location","employment_type","status","salary","date_added"]
        for col,h in enumerate(headers,1):
            cell=ws.cell(row=1,column=col,value=h)
            cell.font=hdr_font; cell.fill=hdr_fill; cell.alignment=Alignment(horizontal="center")
        for row_i,(emp_id,emp) in enumerate(d["employees"].items(),start=2):
            ws.cell(row=row_i,column=1,value=emp_id)
            for col_i,key in enumerate(keys[1:],2):
                ws.cell(row=row_i,column=col_i,value=emp.get(key,""))
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width=18
        wb.save(path)
        messagebox.showinfo("Exported",f"Saved {len(d['employees'])} records to:\n{path}")

    def _logout(self):
        if messagebox.askyesno("Sign Out","Sign out of ARCHER ENTERPRISE?"):
            d = load_data()
            add_audit(d, self.username, "LOGOUT", self.username, "Desktop sign out")
            save_data(d)
            self.destroy()
            LoginWindow().mainloop()


# ════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("\n" + "═"*58)
    print("  🏹  ARCHER ENTERPRISE — Desktop Edition")
    print("  ─"*29)
    print("  Shares archer_data.json with the Web Edition")
    print("═"*58 + "\n")
    LoginWindow().mainloop()
