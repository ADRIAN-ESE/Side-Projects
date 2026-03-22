"""
archer_web.py
══════════════════════════════════════════════════════════════
ARCHER ENTERPRISE — Web Application  (Flask)
Run: python archer_web.py  →  http://localhost:5000

Requirements:
    pip install flask openpyxl

Default login: admin / admin123
══════════════════════════════════════════════════════════════
"""

import os, json, csv, io
from datetime import timedelta, datetime
from functools import wraps
from flask import (Flask, request, jsonify, render_template_string,
                   session, redirect, url_for, send_from_directory, send_file)
import archer_data as db

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

app = Flask(__name__)
app.secret_key = os.environ.get("ARCHER_SECRET", "archer-ent-2024-xK9mP")
app.permanent_session_lifetime = timedelta(hours=8)

ALLOWED_IMG = {".jpg", ".jpeg", ".png", ".gif", ".webp"}

# ═══════════════════════════════════════════════════════════════════════════════
#  AUTH HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def get_current_user():
    uname = session.get("username")
    if not uname:
        return None
    d = db.load_data()
    u = d["auth_users"].get(uname)
    if not u:
        return None
    return {"username": uname, "role": u["role"],
            "full_name": u.get("full_name", ""), "employee_id": u.get("employee_id")}

def login_required(f):
    @wraps(f)
    def decorated(*a, **kw):
        if not get_current_user():
            return redirect(url_for("login_page"))
        return f(*a, **kw)
    return decorated

def api_login_required(f):
    @wraps(f)
    def decorated(*a, **kw):
        if not get_current_user():
            return jsonify({"error": "Unauthorized"}), 401
        return f(*a, **kw)
    return decorated

def perm_required(perm):
    def decorator(f):
        @wraps(f)
        def decorated(*a, **kw):
            u = get_current_user()
            if not u:
                return jsonify({"error": "Unauthorized"}), 401
            if not db.can(u["role"], perm):
                return jsonify({"error": "Forbidden — insufficient permissions"}), 403
            return f(*a, **kw)
        return decorated
    return decorator

# ═══════════════════════════════════════════════════════════════════════════════
#  HTML TEMPLATES
# ═══════════════════════════════════════════════════════════════════════════════

LOGIN_TPL = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>ARCHER ENTERPRISE — Login</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=IBM+Plex+Mono:wght@300;400;500;600&family=Barlow:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#04060d;--card:#090d1a;--border:#16213a;
  --gold:#c9a848;--gold2:#e8c96a;--text:#d5ddf0;--muted:#3a4a6a;
  --err:#ef4444;--bebas:'Bebas Neue',sans-serif;
  --mono:'IBM Plex Mono',monospace;--body:'Barlow',sans-serif;
}
html,body{height:100%;background:var(--bg);font-family:var(--body);color:var(--text);}
body{
  display:flex;align-items:center;justify-content:center;
  background-image:
    radial-gradient(circle at 20% 50%, rgba(201,168,72,.04) 0%, transparent 50%),
    radial-gradient(circle at 80% 20%, rgba(64,128,248,.04) 0%, transparent 40%),
    linear-gradient(rgba(22,33,58,.3) 1px, transparent 1px),
    linear-gradient(90deg, rgba(22,33,58,.3) 1px, transparent 1px);
  background-size:100% 100%,100% 100%,40px 40px,40px 40px;
}
.wrap{width:420px;animation:fadeUp .5s ease both;}
@keyframes fadeUp{from{opacity:0;transform:translateY(24px)}to{opacity:1;transform:none}}
.brand{text-align:center;margin-bottom:36px;}
.brand h1{
  font-family:var(--bebas);font-size:64px;letter-spacing:4px;
  background:linear-gradient(135deg,var(--gold2),var(--gold));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;
  background-clip:text;line-height:1;
}
.brand p{font-size:9px;letter-spacing:3.5px;color:var(--muted);
  text-transform:uppercase;margin-top:6px;font-weight:500;}
.card{
  background:var(--card);border:1px solid var(--border);
  border-top:3px solid var(--gold);border-radius:4px;
  padding:40px 36px;
}
.card h2{font-size:13px;letter-spacing:2px;text-transform:uppercase;
  color:var(--muted);margin-bottom:28px;font-weight:500;}
.field{margin-bottom:20px;}
.field label{display:block;font-size:10px;letter-spacing:1.5px;
  text-transform:uppercase;color:var(--muted);margin-bottom:8px;font-weight:600;}
.field input{
  width:100%;background:#06090f;border:1px solid var(--border);
  color:var(--text);padding:12px 14px;border-radius:3px;
  font-family:var(--mono);font-size:13px;outline:none;
  transition:border-color .2s;
}
.field input:focus{border-color:var(--gold);}
.field input::placeholder{color:var(--muted);}
.btn-login{
  width:100%;padding:14px;background:var(--gold);color:#04060d;
  border:none;border-radius:3px;font-family:var(--bebas);font-size:18px;
  letter-spacing:3px;cursor:pointer;transition:background .2s,transform .1s;
  margin-top:8px;
}
.btn-login:hover{background:var(--gold2);}
.btn-login:active{transform:scale(.98);}
.err{
  background:rgba(239,68,68,.08);border:1px solid rgba(239,68,68,.3);
  color:var(--err);padding:10px 14px;border-radius:3px;
  font-size:12px;margin-bottom:16px;display:none;
}
.err.show{display:block;}
.footer{text-align:center;margin-top:28px;font-size:10px;
  letter-spacing:1.5px;color:var(--muted);text-transform:uppercase;}
</style>
</head>
<body>
<div class="wrap">
  <div class="brand">
    <h1>ARCHER</h1>
    <p>Corporate Management &amp; Service Agency</p>
  </div>
  <div class="card">
    <h2>System Access</h2>
    <div class="err" id="err">{% if error %}{{ error }}{% endif %}</div>
    <form method="POST" action="/login" onsubmit="return validate()">
      <div class="field">
        <label>Username</label>
        <input type="text" name="username" id="uname" placeholder="Enter username" autocomplete="username" required>
      </div>
      <div class="field">
        <label>Password</label>
        <input type="password" name="password" id="pw" placeholder="••••••••" autocomplete="current-password" required>
      </div>
      <button class="btn-login" type="submit">AUTHENTICATE</button>
    </form>
  </div>
  <div class="footer">ARCHER ENTERPRISE &copy; {{ year }}</div>
</div>
<script>
{% if error %} document.getElementById('err').classList.add('show'); {% endif %}
function validate(){
  const u=document.getElementById('uname').value.trim();
  const p=document.getElementById('pw').value;
  if(!u||!p){
    const e=document.getElementById('err');
    e.textContent='Please enter both username and password.';
    e.classList.add('show'); return false;
  }
  return true;
}
</script>
</body></html>"""

APP_TPL = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>ARCHER ENTERPRISE — Management System</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=IBM+Plex+Mono:wght@300;400;500;600&family=Barlow:wght@300;400;500;600&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg:#04060d;--surface:#070a14;--card:#0b0f1e;--card2:#0e1325;
  --border:#141e35;--border2:#1c2840;
  --gold:#c9a848;--gold2:#e8c96a;--gold-dim:rgba(201,168,72,.12);
  --blue:#3b7ef8;--green:#22c55e;--amber:#f59e0b;--red:#ef4444;
  --text:#d0d8ee;--muted:#3c4d6c;--muted2:#5a6e94;
  --sidebar:#040610;--sidebar-w:220px;
  --bebas:'Bebas Neue',sans-serif;
  --mono:'IBM Plex Mono',monospace;
  --body:'Barlow',sans-serif;
  --r:4px;
}
html,body{height:100%;background:var(--bg);color:var(--text);font-family:var(--body);font-size:13px;}
::-webkit-scrollbar{width:5px;height:5px;}
::-webkit-scrollbar-track{background:var(--bg);}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px;}
#app{display:flex;height:100vh;overflow:hidden;}

/* SIDEBAR */
#sidebar{
  width:var(--sidebar-w);flex-shrink:0;background:var(--sidebar);
  border-right:1px solid var(--border);display:flex;flex-direction:column;
  overflow-y:auto;
}
.sb-logo{padding:24px 18px 18px;border-bottom:1px solid var(--border);}
.sb-logo h1{
  font-family:var(--bebas);font-size:28px;letter-spacing:3px;
  background:linear-gradient(135deg,var(--gold2) 0%,var(--gold) 100%);
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;
}
.sb-logo p{font-size:8px;letter-spacing:2px;color:var(--muted);text-transform:uppercase;
  margin-top:3px;line-height:1.4;}
.sb-section{font-size:9px;letter-spacing:2px;text-transform:uppercase;
  color:var(--muted);padding:18px 18px 6px;font-weight:600;}
.nav-item{
  display:flex;align-items:center;gap:10px;padding:10px 18px;
  cursor:pointer;border:none;background:transparent;
  color:var(--muted2);font-family:var(--body);font-size:12.5px;width:100%;
  text-align:left;transition:all .15s;border-left:2px solid transparent;
  text-decoration:none;
}
.nav-item:hover{background:rgba(201,168,72,.04);color:var(--text);}
.nav-item.active{
  background:var(--gold-dim);color:var(--gold);
  border-left-color:var(--gold);
}
.nav-icon{font-size:14px;width:18px;text-align:center;opacity:.8;}
.sb-user{
  margin-top:auto;padding:16px 18px;border-top:1px solid var(--border);
  display:flex;align-items:center;gap:10px;
}
.sb-avatar{
  width:34px;height:34px;border-radius:50%;background:var(--gold-dim);
  border:1px solid var(--gold);display:flex;align-items:center;justify-content:center;
  font-family:var(--bebas);font-size:14px;color:var(--gold);flex-shrink:0;
  overflow:hidden;
}
.sb-avatar img{width:100%;height:100%;object-fit:cover;}
.sb-info{flex:1;min-width:0;}
.sb-name{font-size:12px;font-weight:600;color:var(--text);
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.sb-role{font-size:10px;color:var(--muted2);}
.sb-logout{
  background:none;border:none;cursor:pointer;color:var(--muted);
  font-size:14px;padding:4px;transition:color .15s;
}
.sb-logout:hover{color:var(--red);}

/* MAIN */
#main{flex:1;overflow-y:auto;display:flex;flex-direction:column;}
.topbar{
  background:var(--surface);border-bottom:1px solid var(--border);
  padding:14px 28px;display:flex;align-items:center;gap:12px;flex-shrink:0;
}
.topbar-title{font-family:var(--bebas);font-size:22px;letter-spacing:2px;color:var(--text);}
.topbar-crumb{font-size:10px;letter-spacing:1px;color:var(--muted);text-transform:uppercase;}
.topbar-spacer{flex:1;}
.topbar-badge{
  background:var(--gold-dim);border:1px solid var(--gold);
  color:var(--gold);font-size:9px;letter-spacing:1.5px;text-transform:uppercase;
  padding:4px 10px;border-radius:2px;font-weight:600;
}
.content{padding:24px 28px;flex:1;}
.page{display:none;}
.page.active{display:block;}

/* STAT CARDS */
.stat-row{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:24px;}
.stat-card{
  background:var(--card);border:1px solid var(--border);border-radius:var(--r);
  padding:20px 18px;position:relative;overflow:hidden;
}
.stat-card::before{
  content:'';position:absolute;top:0;left:0;right:0;height:2px;
  background:var(--gold);
}
.stat-card.blue::before{background:var(--blue);}
.stat-card.green::before{background:var(--green);}
.stat-card.amber::before{background:var(--amber);}
.stat-val{font-family:var(--bebas);font-size:42px;color:var(--gold);letter-spacing:2px;line-height:1;}
.stat-card.blue .stat-val{color:var(--blue);}
.stat-card.green .stat-val{color:var(--green);}
.stat-card.amber .stat-val{color:var(--amber);}
.stat-label{font-size:10px;letter-spacing:1.5px;text-transform:uppercase;
  color:var(--muted2);margin-top:4px;font-weight:500;}

/* CHARTS */
.chart-row{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:24px;}
.chart-card{background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:20px;}
.chart-title{font-size:10px;letter-spacing:2px;text-transform:uppercase;
  color:var(--muted2);margin-bottom:16px;font-weight:600;}
.chart-wrap{position:relative;height:200px;}

/* CARDS & SECTION HEADERS */
.sec-header{display:flex;align-items:center;gap:12px;margin-bottom:18px;flex-wrap:wrap;}
.sec-title{font-family:var(--bebas);font-size:20px;letter-spacing:2px;}
.sec-spacer{flex:1;}
.card{background:var(--card);border:1px solid var(--border);border-radius:var(--r);padding:22px;}

/* TOOLBAR */
.toolbar{display:flex;align-items:center;gap:10px;margin-bottom:14px;flex-wrap:wrap;}
.search-inp{
  flex:1;min-width:220px;max-width:380px;
  background:var(--surface);border:1px solid var(--border);
  color:var(--text);padding:9px 13px;border-radius:var(--r);
  font-family:var(--mono);font-size:12px;outline:none;transition:border-color .15s;
}
.search-inp:focus{border-color:var(--gold);}
.search-inp::placeholder{color:var(--muted);}
.t-spacer{flex:1;}

/* BUTTONS */
.btn{
  display:inline-flex;align-items:center;gap:7px;
  padding:8px 16px;border-radius:var(--r);border:none;
  font-family:var(--body);font-size:12.5px;font-weight:600;
  cursor:pointer;transition:all .15s;white-space:nowrap;letter-spacing:.3px;
}
.btn-gold{background:var(--gold);color:#04060d;}
.btn-gold:hover{background:var(--gold2);}
.btn-outline{background:transparent;border:1px solid var(--border2);color:var(--muted2);}
.btn-outline:hover{border-color:var(--gold);color:var(--gold);}
.btn-danger{background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.3);color:var(--red);}
.btn-danger:hover{background:rgba(239,68,68,.2);}
.btn-success{background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.3);color:var(--green);}
.btn-success:hover{background:rgba(34,197,94,.2);}
.btn-sm{padding:5px 10px;font-size:11px;}

/* TABLE */
.tbl-wrap{background:var(--card);border:1px solid var(--border);border-radius:var(--r);overflow:auto;}
table{width:100%;border-collapse:collapse;}
thead th{
  background:var(--surface);color:var(--muted2);font-size:9.5px;font-weight:700;
  letter-spacing:1.5px;text-transform:uppercase;padding:11px 14px;
  text-align:left;white-space:nowrap;border-bottom:1px solid var(--border);
  position:sticky;top:0;z-index:1;
}
tbody tr{border-bottom:1px solid rgba(20,30,53,.8);transition:background .1s;cursor:pointer;}
tbody tr:hover{background:rgba(201,168,72,.04);}
tbody tr.selected{background:rgba(201,168,72,.07);}
tbody tr:last-child{border-bottom:none;}
td{padding:10px 14px;font-family:var(--mono);font-size:11.5px;white-space:nowrap;}
td.name-col{font-family:var(--body);font-size:12.5px;font-weight:500;}
.empty-row td{text-align:center;color:var(--muted);padding:36px;
  font-family:var(--body);font-size:13px;}

/* BADGES */
.badge{display:inline-block;padding:2px 8px;border-radius:2px;
  font-size:10px;font-weight:700;letter-spacing:.5px;text-transform:uppercase;}
.badge-active{background:rgba(34,197,94,.1);color:var(--green);border:1px solid rgba(34,197,94,.2);}
.badge-inactive{background:rgba(90,110,148,.1);color:var(--muted2);border:1px solid var(--border);}
.badge-leave{background:rgba(245,158,11,.1);color:var(--amber);border:1px solid rgba(245,158,11,.2);}
.badge-terminated{background:rgba(239,68,68,.1);color:var(--red);border:1px solid rgba(239,68,68,.2);}
.badge-pending{background:rgba(245,158,11,.1);color:var(--amber);border:1px solid rgba(245,158,11,.2);}
.badge-approved{background:rgba(34,197,94,.1);color:var(--green);border:1px solid rgba(34,197,94,.2);}
.badge-denied{background:rgba(239,68,68,.1);color:var(--red);border:1px solid rgba(239,68,68,.2);}
.badge-admin{background:rgba(201,168,72,.1);color:var(--gold);border:1px solid rgba(201,168,72,.2);}
.badge-hr{background:rgba(59,126,248,.1);color:var(--blue);border:1px solid rgba(59,126,248,.2);}
.badge-manager{background:rgba(168,85,247,.1);color:#c084fc;border:1px solid rgba(168,85,247,.2);}
.badge-staff{background:rgba(90,110,148,.1);color:var(--muted2);border:1px solid var(--border);}

/* AVATAR CELL */
.av-cell{display:flex;align-items:center;gap:10px;}
.av-mini{
  width:30px;height:30px;border-radius:50%;flex-shrink:0;
  background:var(--gold-dim);border:1px solid rgba(201,168,72,.3);
  display:flex;align-items:center;justify-content:center;
  font-family:var(--bebas);font-size:12px;color:var(--gold);overflow:hidden;
}
.av-mini img{width:100%;height:100%;object-fit:cover;}

/* DEPT GRID */
.dept-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:14px;}
.dept-card{
  background:var(--card);border:1px solid var(--border);border-radius:var(--r);
  padding:20px;transition:border-color .2s;
}
.dept-card:hover{border-color:var(--gold);}
.dept-card-name{font-family:var(--bebas);font-size:20px;letter-spacing:1.5px;color:var(--gold);margin-bottom:4px;}
.dept-card-head{font-size:11px;color:var(--muted2);margin-bottom:8px;}
.dept-card-desc{font-size:12px;color:var(--muted2);margin-bottom:14px;line-height:1.5;}
.dept-card-count{font-size:11px;letter-spacing:1px;text-transform:uppercase;color:var(--muted);}
.dept-card-count span{color:var(--text);font-weight:600;}
.dept-actions{display:flex;gap:6px;margin-top:14px;}

/* MODAL */
.modal-overlay{
  position:fixed;inset:0;background:rgba(0,0,0,.75);
  display:flex;align-items:center;justify-content:center;z-index:100;
  backdrop-filter:blur(3px);opacity:0;pointer-events:none;transition:opacity .2s;
}
.modal-overlay.open{opacity:1;pointer-events:all;}
.modal{
  background:var(--card2);border:1px solid var(--border2);border-top:2px solid var(--gold);
  border-radius:var(--r);padding:0;width:580px;max-width:95vw;
  max-height:90vh;overflow-y:auto;
  transform:translateY(16px);transition:transform .2s;
}
.modal-overlay.open .modal{transform:translateY(0);}
.modal-header{padding:20px 24px 0;display:flex;align-items:center;justify-content:space-between;}
.modal-title{font-family:var(--bebas);font-size:20px;letter-spacing:2px;}
.modal-close{background:none;border:none;color:var(--muted);cursor:pointer;font-size:18px;
  transition:color .15s;padding:4px;}
.modal-close:hover{color:var(--text);}
.modal-body{padding:20px 24px 24px;}

/* FORM */
.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;}
.form-group{display:flex;flex-direction:column;gap:6px;}
.form-group.full{grid-column:1/-1;}
.form-group label{font-size:9.5px;letter-spacing:1.5px;text-transform:uppercase;
  color:var(--muted2);font-weight:600;}
.form-group input,.form-group select,.form-group textarea{
  background:#06090f;border:1px solid var(--border);color:var(--text);
  padding:9px 12px;border-radius:var(--r);font-family:var(--mono);font-size:12px;
  outline:none;transition:border-color .15s;
}
.form-group input:focus,.form-group select:focus,.form-group textarea:focus{border-color:var(--gold);}
.form-group input::placeholder{color:var(--muted);}
.form-group select option{background:#06090f;}
.modal-actions{display:flex;gap:10px;justify-content:flex-end;margin-top:20px;}

/* AVATAR UPLOAD */
.av-upload-zone{
  border:2px dashed var(--border2);border-radius:var(--r);
  padding:16px;text-align:center;cursor:pointer;transition:border-color .2s;
}
.av-upload-zone:hover{border-color:var(--gold);}
.av-preview{
  width:72px;height:72px;border-radius:50%;
  background:var(--gold-dim);border:2px solid var(--gold);
  margin:0 auto 10px;overflow:hidden;display:flex;
  align-items:center;justify-content:center;
  font-family:var(--bebas);font-size:24px;color:var(--gold);
}
.av-preview img{width:100%;height:100%;object-fit:cover;}
.av-upload-label{font-size:11px;color:var(--muted2);}

/* AUDIT TABLE */
.audit-action{
  display:inline-block;padding:2px 8px;border-radius:2px;
  font-size:10px;font-weight:700;letter-spacing:.4px;text-transform:uppercase;
  background:rgba(59,126,248,.1);color:var(--blue);border:1px solid rgba(59,126,248,.2);
}

/* TOAST */
#toast{
  position:fixed;bottom:24px;right:24px;z-index:200;
  background:var(--card2);border:1px solid var(--border2);border-left:3px solid var(--gold);
  padding:13px 18px;border-radius:var(--r);font-size:13px;max-width:320px;
  transform:translateY(60px);opacity:0;transition:all .3s;
}
#toast.show{transform:translateY(0);opacity:1;}
#toast.err{border-left-color:var(--red);}
#toast.ok{border-left-color:var(--green);}

/* RECENT ACTIVITY */
.activity-row{display:flex;gap:14px;padding:10px 0;
  border-bottom:1px solid rgba(20,30,53,.8);align-items:flex-start;}
.activity-row:last-child{border-bottom:none;}
.act-dot{width:7px;height:7px;border-radius:50%;background:var(--gold);
  flex-shrink:0;margin-top:5px;}
.act-text{flex:1;font-size:12px;line-height:1.5;}
.act-time{font-size:10.5px;color:var(--muted);white-space:nowrap;font-family:var(--mono);}

/* IO grid */
.io-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;}

/* RESPONSIVE */
@media(max-width:900px){
  .stat-row{grid-template-columns:1fr 1fr;}
  .chart-row{grid-template-columns:1fr;}
  .form-grid{grid-template-columns:1fr;}
  .io-grid{grid-template-columns:1fr;}
}
</style>
</head>
<body>
<script>const USER = {{ user_json | safe }};</script>
<div id="app">

<!-- SIDEBAR -->
<nav id="sidebar">
  <div class="sb-logo">
    <h1>ARCHER</h1>
    <p>Corporate Management<br>&amp; Service Agency</p>
  </div>
  <div class="sb-section">Navigation</div>
  <button class="nav-item active" data-tab="dashboard"><span class="nav-icon">▦</span> Dashboard</button>
  <button class="nav-item" data-tab="employees"><span class="nav-icon">◈</span> Employees</button>
  <button class="nav-item" data-tab="departments"><span class="nav-icon">◉</span> Departments</button>
  <button class="nav-item" data-tab="promotions"><span class="nav-icon">△</span> Promotions</button>
  <div id="sb-audit-wrap">
  <button class="nav-item" data-tab="audit"><span class="nav-icon">≡</span> Audit Log</button>
  </div>
  <div id="sb-users-wrap" style="display:none">
  <div class="sb-section">Admin</div>
  <button class="nav-item" data-tab="users"><span class="nav-icon">⊕</span> System Users</button>
  </div>
  <div class="sb-user">
    <div class="sb-avatar" id="sb-av">?</div>
    <div class="sb-info">
      <div class="sb-name" id="sb-name">—</div>
      <div class="sb-role" id="sb-role">—</div>
    </div>
    <button class="sb-logout" onclick="window.location='/logout'" title="Logout">⏻</button>
  </div>
</nav>

<!-- MAIN -->
<div id="main">
  <div class="topbar">
    <div class="topbar-title" id="tb-title">DASHBOARD</div>
    <span class="topbar-crumb" id="tb-crumb">ARCHER ENTERPRISE / DASHBOARD</span>
    <div class="topbar-spacer"></div>
    <span class="topbar-badge" id="tb-role-badge">—</span>
  </div>
  <div class="content">

    <!-- ── DASHBOARD ────────────────────────────────────────────────────── -->
    <section class="page active" id="page-dashboard">
      <div class="stat-row">
        <div class="stat-card"><div class="stat-val" id="s-total">0</div><div class="stat-label">Total Employees</div></div>
        <div class="stat-card green"><div class="stat-val" id="s-active">0</div><div class="stat-label">Active Staff</div></div>
        <div class="stat-card blue"><div class="stat-val" id="s-depts">0</div><div class="stat-label">Departments</div></div>
        <div class="stat-card amber"><div class="stat-val" id="s-promo">0</div><div class="stat-label">Pending Promotions</div></div>
      </div>
      <div class="chart-row">
        <div class="chart-card">
          <div class="chart-title">Employees by Department</div>
          <div class="chart-wrap"><canvas id="chartDept"></canvas></div>
        </div>
        <div class="chart-card">
          <div class="chart-title">Employment Status</div>
          <div class="chart-wrap"><canvas id="chartStatus"></canvas></div>
        </div>
      </div>
      <div class="card">
        <div class="chart-title" style="margin-bottom:12px">Recent Activity</div>
        <div id="activity-feed"></div>
      </div>
    </section>

    <!-- ── EMPLOYEES ────────────────────────────────────────────────────── -->
    <section class="page" id="page-employees">
      <div class="sec-header">
        <div class="sec-title">EMPLOYEES</div>
        <input class="search-inp" id="emp-search" placeholder="Search name, ID, department, role…" oninput="filterEmployees()">
        <div class="sec-spacer"></div>
        <button class="btn btn-outline btn-sm" onclick="exportData('csv')">↓ CSV</button>
        <button class="btn btn-outline btn-sm" onclick="exportData('excel')">↓ Excel</button>
        <button class="btn btn-gold" id="btn-add-emp" onclick="openAddEmpModal()">＋ Add Employee</button>
      </div>
      <div class="tbl-wrap">
        <table>
          <thead>
            <tr>
              <th>Employee ID</th><th>Name</th><th>Department</th>
              <th>Position / Level</th><th>Status</th><th>Joined</th><th>Actions</th>
            </tr>
          </thead>
          <tbody id="emp-tbody"></tbody>
        </table>
      </div>
    </section>

    <!-- ── DEPARTMENTS ───────────────────────────────────────────────────── -->
    <section class="page" id="page-departments">
      <div class="sec-header">
        <div class="sec-title">DEPARTMENTS</div>
        <div class="sec-spacer"></div>
        <button class="btn btn-gold" id="btn-add-dept" onclick="openAddDeptModal()">＋ Add Department</button>
      </div>
      <div class="dept-grid" id="dept-grid"></div>
    </section>

    <!-- ── PROMOTIONS ────────────────────────────────────────────────────── -->
    <section class="page" id="page-promotions">
      <div class="sec-header"><div class="sec-title">PROMOTIONS</div></div>
      <div class="card" style="margin-bottom:18px">
        <div class="chart-title" style="margin-bottom:14px">Submit Promotion Request</div>
        <div class="form-grid">
          <div class="form-group">
            <label>Employee ID or Name</label>
            <input id="p-emp" placeholder="ARC-2024-0001 or search name">
          </div>
          <div class="form-group">
            <label>Requested Level / Title</label>
            <input id="p-role" placeholder="e.g. Senior Engineer, Director">
          </div>
        </div>
        <div style="margin-top:12px">
          <button class="btn btn-gold" onclick="submitPromo()">Submit Request</button>
        </div>
      </div>
      <div class="sec-header" style="margin-bottom:10px">
        <div class="sec-title" style="font-size:16px">ALL REQUESTS</div>
        <div class="sec-spacer"></div>
        <button class="btn btn-success btn-sm" id="btn-approve" onclick="resolvePromo('approved')">✓ Approve</button>
        <button class="btn btn-danger btn-sm"  id="btn-deny"    onclick="resolvePromo('denied')">✕ Deny</button>
      </div>
      <div class="tbl-wrap">
        <table>
          <thead>
            <tr><th>Employee</th><th>Current Level</th><th>Requested</th><th>Submitted</th><th>Status</th></tr>
          </thead>
          <tbody id="promo-tbody"></tbody>
        </table>
      </div>
    </section>

    <!-- ── AUDIT LOG ──────────────────────────────────────────────────────── -->
    <section class="page" id="page-audit">
      <div class="sec-header">
        <div class="sec-title">AUDIT LOG</div>
        <input class="search-inp" id="audit-search" placeholder="Filter by action, user, or target…" oninput="filterAudit()">
      </div>
      <div class="tbl-wrap">
        <table>
          <thead>
            <tr><th>Timestamp</th><th>Action</th><th>Performed By</th><th>Target</th><th>Details</th></tr>
          </thead>
          <tbody id="audit-tbody"></tbody>
        </table>
      </div>
    </section>

    <!-- ── SYSTEM USERS ───────────────────────────────────────────────────── -->
    <section class="page" id="page-users">
      <div class="sec-header">
        <div class="sec-title">SYSTEM USERS</div>
        <div class="sec-spacer"></div>
        <button class="btn btn-gold" onclick="openAddUserModal()">＋ Add User</button>
      </div>
      <div class="tbl-wrap">
        <table>
          <thead>
            <tr><th>Username</th><th>Full Name</th><th>Role</th><th>Linked Employee</th><th>Created</th><th>Actions</th></tr>
          </thead>
          <tbody id="users-tbody"></tbody>
        </table>
      </div>
    </section>

  </div><!-- /content -->
</div><!-- /main -->
</div><!-- /app -->

<!-- ── MODALS ──────────────────────────────────────────────────────────────── -->

<!-- Add/Edit Employee -->
<div class="modal-overlay" id="emp-modal">
<div class="modal">
  <div class="modal-header">
    <div class="modal-title" id="emp-modal-title">ADD EMPLOYEE</div>
    <button class="modal-close" onclick="closeModal('emp-modal')">✕</button>
  </div>
  <div class="modal-body">
    <input type="hidden" id="em-id">
    <div style="margin-bottom:16px">
      <div class="av-upload-zone" onclick="document.getElementById('av-file').click()">
        <div class="av-preview" id="av-preview-wrap"><span id="av-initials">?</span></div>
        <div class="av-upload-label">Click to upload photo (JPG, PNG)</div>
        <input type="file" id="av-file" accept=".jpg,.jpeg,.png,.webp" style="display:none" onchange="previewAvatar(this)">
      </div>
    </div>
    <div class="form-grid">
      <div class="form-group"><label>Full Name *</label><input id="em-name" placeholder="First Last"></div>
      <div class="form-group"><label>Email</label><input id="em-email" type="email" placeholder="name@company.com"></div>
      <div class="form-group"><label>Phone</label><input id="em-phone" placeholder="+1 000 000 0000"></div>
      <div class="form-group"><label>Age</label><input id="em-age" placeholder="e.g. 30"></div>
      <div class="form-group"><label>Department</label><select id="em-dept"></select></div>
      <div class="form-group"><label>Position / Title</label><input id="em-position" placeholder="e.g. Software Engineer"></div>
      <div class="form-group"><label>Level</label><select id="em-level"></select></div>
      <div class="form-group"><label>Status</label><select id="em-status"></select></div>
      <div class="form-group full"><label>Location</label><input id="em-location" placeholder="City, Country"></div>
    </div>
    <div class="modal-actions">
      <button class="btn btn-outline" onclick="closeModal('emp-modal')">Cancel</button>
      <button class="btn btn-gold" onclick="saveEmployee()">Save Employee</button>
    </div>
  </div>
</div></div>

<!-- Add Department -->
<div class="modal-overlay" id="dept-modal">
<div class="modal">
  <div class="modal-header">
    <div class="modal-title" id="dept-modal-title">ADD DEPARTMENT</div>
    <button class="modal-close" onclick="closeModal('dept-modal')">✕</button>
  </div>
  <div class="modal-body">
    <input type="hidden" id="dm-id">
    <div class="form-grid">
      <div class="form-group"><label>Department Name *</label><input id="dm-name" placeholder="e.g. Engineering"></div>
      <div class="form-group"><label>Department Head</label><input id="dm-head" placeholder="Employee name or ID"></div>
      <div class="form-group full"><label>Description</label><input id="dm-desc" placeholder="Brief description"></div>
    </div>
    <div class="modal-actions">
      <button class="btn btn-outline" onclick="closeModal('dept-modal')">Cancel</button>
      <button class="btn btn-gold" onclick="saveDept()">Save</button>
    </div>
  </div>
</div></div>

<!-- Add System User -->
<div class="modal-overlay" id="user-modal">
<div class="modal" style="width:440px">
  <div class="modal-header">
    <div class="modal-title">ADD SYSTEM USER</div>
    <button class="modal-close" onclick="closeModal('user-modal')">✕</button>
  </div>
  <div class="modal-body">
    <div class="form-grid">
      <div class="form-group"><label>Username *</label><input id="um-uname" placeholder="login name"></div>
      <div class="form-group"><label>Full Name</label><input id="um-name" placeholder="First Last"></div>
      <div class="form-group"><label>Password *</label><input id="um-pw" type="password" placeholder="••••••••"></div>
      <div class="form-group"><label>Role</label><select id="um-role"></select></div>
    </div>
    <div class="modal-actions">
      <button class="btn btn-outline" onclick="closeModal('user-modal')">Cancel</button>
      <button class="btn btn-gold" onclick="saveUser()">Create User</button>
    </div>
  </div>
</div></div>

<div id="toast"></div>

{% raw %}
<script>
// ── State ──────────────────────────────────────────────────────────────────
let employees={}, departments={}, allPromos=[], auditLog=[], systemUsers={};
let selEmpId=null, selPromoKey=null;
let deptChart=null, statusChart=null;

// ── Init ───────────────────────────────────────────────────────────────────
(async()=>{
  setupNav();
  setupUI();
  await Promise.all([loadEmployees(),loadDepts(),loadPromos(),loadAudit(),loadUsers()]);
  loadDashboard();
})();

// ── Sidebar ────────────────────────────────────────────────────────────────
function setupNav(){
  const role=USER.role;
  document.getElementById('sb-name').textContent=USER.full_name||USER.username;
  document.getElementById('sb-role').textContent=role;
  document.getElementById('tb-role-badge').textContent=role;
  const av=document.getElementById('sb-av');
  const n=(USER.full_name||USER.username);
  av.textContent=n.split(' ').map(x=>x[0]).join('').slice(0,2).toUpperCase();

  if(role==='Admin'){
    document.getElementById('sb-users-wrap').style.display='block';
  }
  if(!['Admin','HR'].includes(role)){
    const aw=document.getElementById('sb-audit-wrap');
    if(aw) aw.style.display='none';
  }

  document.querySelectorAll('.nav-item').forEach(btn=>{
    btn.addEventListener('click',()=>{
      const tab=btn.dataset.tab;
      document.querySelectorAll('.nav-item').forEach(b=>b.classList.remove('active'));
      document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
      btn.classList.add('active');
      document.getElementById('page-'+tab).classList.add('active');
      const titles={dashboard:'DASHBOARD',employees:'EMPLOYEES',departments:'DEPARTMENTS',
                    promotions:'PROMOTIONS',audit:'AUDIT LOG',users:'SYSTEM USERS'};
      document.getElementById('tb-title').textContent=titles[tab]||tab.toUpperCase();
      document.getElementById('tb-crumb').textContent=`ARCHER ENTERPRISE / ${(titles[tab]||tab.toUpperCase())}`;
      if(tab==='dashboard') loadDashboard();
      if(tab==='employees') renderEmployees();
      if(tab==='departments') renderDepts();
      if(tab==='promotions') renderPromos();
      if(tab==='audit') renderAudit();
      if(tab==='users') renderUsers();
    });
  });
}

function setupUI(){
  // Populate select options
  const levels=['Intern','Junior','Mid-Level','Senior','Lead','Manager','Director','C-Level'];
  const statuses=['Active','On Leave','Inactive','Terminated'];
  const roles=['Admin','HR','Manager','Staff'];
  populate('em-level',levels);
  populate('em-status',statuses);
  populate('um-role',roles);

  // Permission-based UI
  const role=USER.role;
  if(!['Admin','HR','Manager'].includes(role)){
    const b=document.getElementById('btn-add-emp');
    if(b) b.style.display='none';
  }
  if(!['Admin','HR'].includes(role)){
    const b=document.getElementById('btn-add-dept');
    if(b) b.style.display='none';
    document.getElementById('btn-approve').style.display='none';
    document.getElementById('btn-deny').style.display='none';
  }
}

function populate(selectId, options){
  const s=document.getElementById(selectId);
  if(!s) return;
  s.innerHTML=options.map(o=>`<option value="${esc(o)}">${esc(o)}</option>`).join('');
}

// ── DATA LOADERS ───────────────────────────────────────────────────────────
async function loadEmployees(){ const r=await fetch('/api/employees'); employees=await r.json(); }
async function loadDepts(){ const r=await fetch('/api/departments'); departments=await r.json(); }
async function loadPromos(){ const r=await fetch('/api/promotions'); allPromos=await r.json(); }
async function loadAudit(){ const r=await fetch('/api/audit'); auditLog=await r.json(); }
async function loadUsers(){
  if(USER.role!=='Admin') return;
  const r=await fetch('/api/auth-users'); systemUsers=await r.json();
}

// ── DASHBOARD ──────────────────────────────────────────────────────────────
async function loadDashboard(){
  const r=await fetch('/api/stats');
  const s=await r.json();
  document.getElementById('s-total').textContent=s.total;
  document.getElementById('s-active').textContent=s.active;
  document.getElementById('s-depts').textContent=s.departments;
  document.getElementById('s-promo').textContent=s.pending_promotions;
  renderDeptChart(s.by_dept);
  renderStatusChart(s.by_status);
  renderActivity();
}

function renderDeptChart(data){
  const labels=Object.keys(data);
  const vals=Object.values(data);
  const ctx=document.getElementById('chartDept');
  if(deptChart) deptChart.destroy();
  deptChart=new Chart(ctx,{
    type:'bar',
    data:{labels,datasets:[{data:vals,backgroundColor:'rgba(201,168,72,.6)',
      borderColor:'#c9a848',borderWidth:1,borderRadius:2}]},
    options:{
      indexAxis:'y',responsive:true,maintainAspectRatio:false,
      plugins:{legend:{display:false}},
      scales:{
        x:{ticks:{color:'#3c4d6c',font:{family:'IBM Plex Mono',size:11}},
           grid:{color:'rgba(20,30,53,.8)'},border:{color:'#141e35'}},
        y:{ticks:{color:'#5a6e94',font:{family:'IBM Plex Mono',size:11}},
           grid:{display:false},border:{color:'#141e35'}}
      }
    }
  });
}

function renderStatusChart(data){
  const labels=Object.keys(data);
  const vals=Object.values(data);
  const colors={'Active':'#22c55e','On Leave':'#f59e0b','Inactive':'#4a5778','Terminated':'#ef4444'};
  const ctx=document.getElementById('chartStatus');
  if(statusChart) statusChart.destroy();
  statusChart=new Chart(ctx,{
    type:'doughnut',
    data:{labels,datasets:[{data:vals,
      backgroundColor:labels.map(l=>colors[l]||'#3b7ef8'),
      borderColor:'#0b0f1e',borderWidth:3}]},
    options:{
      responsive:true,maintainAspectRatio:false,cutout:'65%',
      plugins:{legend:{position:'right',labels:{color:'#5a6e94',
        font:{family:'IBM Plex Mono',size:11},boxWidth:12,padding:12}}}
    }
  });
}

function renderActivity(){
  const feed=document.getElementById('activity-feed');
  const items=auditLog.slice(0,8);
  if(!items.length){feed.innerHTML='<div style="color:var(--muted);font-size:12px;padding:12px 0">No activity yet.</div>';return;}
  feed.innerHTML=items.map(a=>`
    <div class="activity-row">
      <div class="act-dot"></div>
      <div class="act-text"><strong>${esc(a.performed_by)}</strong> — ${esc(a.action)}
        ${a.target?`<span style="color:var(--muted2)"> → ${esc(a.target)}</span>`:''}
        ${a.details?`<br><span style="color:var(--muted);font-size:11px">${esc(a.details)}</span>`:''}
      </div>
      <div class="act-time">${esc(a.timestamp)}</div>
    </div>`).join('');
}

// ── EMPLOYEES ──────────────────────────────────────────────────────────────
function renderEmployees(){
  const q=(document.getElementById('emp-search')?.value||'').toLowerCase();
  const tbody=document.getElementById('emp-tbody');
  const list=Object.entries(employees).filter(([id,e])=>!q||
    [id,e.full_name,e.department,e.position,e.level,e.status,e.email]
    .some(v=>(v||'').toLowerCase().includes(q)));
  if(!list.length){
    tbody.innerHTML='<tr class="empty-row"><td colspan="7">No employees found.</td></tr>';
    return;
  }
  tbody.innerHTML=list.map(([id,e])=>`
    <tr data-id="${esc(id)}" onclick="selectEmp('${esc(id)}',this)">
      <td><span style="font-family:var(--mono);color:var(--gold)">${esc(id)}</span></td>
      <td class="name-col">
        <div class="av-cell">
          ${avatarHtml(e,30)}
          <span>${esc(e.full_name||'—')}</span>
        </div>
      </td>
      <td>${esc(e.department||'—')}</td>
      <td>${esc(e.position||'—')} <span style="color:var(--muted);font-size:10px">/ ${esc(e.level||'')}</span></td>
      <td>${statusBadge(e.status)}</td>
      <td>${esc(e.date_joined||'—')}</td>
      <td>
        <button class="btn btn-outline btn-sm" onclick="event.stopPropagation();editEmp('${esc(id)}')">✎ Edit</button>
        <button class="btn btn-danger btn-sm" onclick="event.stopPropagation();deleteEmp('${esc(id)}')">✕</button>
      </td>
    </tr>`).join('');
}
function filterEmployees(){ renderEmployees(); }
function selectEmp(id,row){
  document.querySelectorAll('#emp-tbody tr').forEach(r=>r.classList.remove('selected'));
  row.classList.add('selected'); selEmpId=id;
}

// ── EMPLOYEE MODAL ─────────────────────────────────────────────────────────
function openAddEmpModal(){
  document.getElementById('emp-modal-title').textContent='ADD EMPLOYEE';
  document.getElementById('em-id').value='';
  ['em-name','em-email','em-phone','em-age','em-position','em-location']
    .forEach(id=>document.getElementById(id).value='');
  document.getElementById('em-dept').innerHTML=
    '<option value="">— Select —</option>'+
    Object.values(departments).map(d=>`<option>${esc(d.name)}</option>`).join('');
  document.getElementById('em-level').value='Junior';
  document.getElementById('em-status').value='Active';
  document.getElementById('av-initials').textContent='?';
  document.getElementById('av-preview-wrap').innerHTML='<span id="av-initials">?</span>';
  document.getElementById('av-file').value='';
  openModal('emp-modal');
}

function editEmp(id){
  const e=employees[id];
  if(!e) return;
  document.getElementById('emp-modal-title').textContent='EDIT EMPLOYEE';
  document.getElementById('em-id').value=id;
  document.getElementById('em-name').value=e.full_name||'';
  document.getElementById('em-email').value=e.email||'';
  document.getElementById('em-phone').value=e.phone||'';
  document.getElementById('em-age').value=e.age||'';
  document.getElementById('em-position').value=e.position||'';
  document.getElementById('em-location').value=e.location||'';
  document.getElementById('em-dept').innerHTML=
    '<option value="">— Select —</option>'+
    Object.values(departments).map(d=>`<option ${d.name===e.department?'selected':''}>${esc(d.name)}</option>`).join('');
  document.getElementById('em-level').value=e.level||'Junior';
  document.getElementById('em-status').value=e.status||'Active';
  const pw=document.getElementById('av-preview-wrap');
  if(e.avatar){
    pw.innerHTML=`<img src="/avatars/${esc(e.avatar)}" alt="avatar">`;
  } else {
    const init=(e.full_name||'?').split(' ').map(x=>x[0]).join('').slice(0,2).toUpperCase();
    pw.innerHTML=`<span id="av-initials">${init}</span>`;
  }
  openModal('emp-modal');
}

async function saveEmployee(){
  const id=document.getElementById('em-id').value;
  const name=document.getElementById('em-name').value.trim();
  if(!name){toast('Full name is required.','err');return;}
  const payload={
    full_name:name, email:document.getElementById('em-email').value.trim(),
    phone:document.getElementById('em-phone').value.trim(),
    age:document.getElementById('em-age').value.trim(),
    department:document.getElementById('em-dept').value,
    position:document.getElementById('em-position').value.trim(),
    level:document.getElementById('em-level').value,
    status:document.getElementById('em-status').value,
    location:document.getElementById('em-location').value.trim(),
  };
  const url=id?`/api/employees/${encodeURIComponent(id)}`:'/api/employees';
  const method=id?'PUT':'POST';
  const r=await fetch(url,{method,headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
  const data=await r.json();
  if(!r.ok){toast(data.error,'err');return;}
  const empId=data.employee_id||id;
  // Avatar upload
  const file=document.getElementById('av-file').files[0];
  if(file&&empId){
    const fd=new FormData(); fd.append('avatar',file);
    await fetch(`/api/employees/${encodeURIComponent(empId)}/avatar`,{method:'POST',body:fd});
  }
  await loadEmployees();
  await loadAudit();
  closeModal('emp-modal');
  renderEmployees();
  toast(id?'Employee updated.':'Employee added.');
  loadDashboard();
}

async function deleteEmp(id){
  if(!confirm(`Delete employee ${id}? This cannot be undone.`)) return;
  const r=await fetch(`/api/employees/${encodeURIComponent(id)}`,{method:'DELETE'});
  if(!r.ok){const d=await r.json();toast(d.error,'err');return;}
  await loadEmployees(); await loadAudit();
  renderEmployees(); toast('Employee deleted.'); loadDashboard();
}

function previewAvatar(input){
  const file=input.files[0]; if(!file) return;
  const reader=new FileReader();
  reader.onload=e=>{
    document.getElementById('av-preview-wrap').innerHTML=
      `<img src="${e.target.result}" alt="preview">`;
  };
  reader.readAsDataURL(file);
}

// ── DEPARTMENTS ────────────────────────────────────────────────────────────
function renderDepts(){
  const grid=document.getElementById('dept-grid');
  const list=Object.entries(departments);
  if(!list.length){grid.innerHTML='<div style="color:var(--muted)">No departments yet.</div>';return;}
  const empsByDept={};
  Object.values(employees).forEach(e=>{
    const d=e.department||'Unassigned';
    empsByDept[d]=(empsByDept[d]||0)+1;
  });
  grid.innerHTML=list.map(([id,d])=>`
    <div class="dept-card">
      <div class="dept-card-name">${esc(d.name)}</div>
      <div class="dept-card-head">Head: ${esc(d.head||'—')}</div>
      <div class="dept-card-desc">${esc(d.description||'—')}</div>
      <div class="dept-card-count">Employees: <span>${empsByDept[d.name]||0}</span></div>
      <div class="dept-actions">
        <button class="btn btn-outline btn-sm" onclick="editDept('${esc(id)}')">✎ Edit</button>
        <button class="btn btn-danger btn-sm" onclick="deleteDept('${esc(id)}')">✕ Delete</button>
      </div>
    </div>`).join('');
}

function openAddDeptModal(){
  document.getElementById('dept-modal-title').textContent='ADD DEPARTMENT';
  document.getElementById('dm-id').value='';
  ['dm-name','dm-head','dm-desc'].forEach(id=>document.getElementById(id).value='');
  openModal('dept-modal');
}
function editDept(id){
  const d=departments[id]; if(!d) return;
  document.getElementById('dept-modal-title').textContent='EDIT DEPARTMENT';
  document.getElementById('dm-id').value=id;
  document.getElementById('dm-name').value=d.name||'';
  document.getElementById('dm-head').value=d.head||'';
  document.getElementById('dm-desc').value=d.description||'';
  openModal('dept-modal');
}
async function saveDept(){
  const id=document.getElementById('dm-id').value;
  const name=document.getElementById('dm-name').value.trim();
  if(!name){toast('Department name is required.','err');return;}
  const payload={name,head:document.getElementById('dm-head').value.trim(),
                 description:document.getElementById('dm-desc').value.trim()};
  const url=id?`/api/departments/${encodeURIComponent(id)}`:'/api/departments';
  const r=await fetch(url,{method:id?'PUT':'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
  if(!r.ok){const d=await r.json();toast(d.error,'err');return;}
  await loadDepts(); closeModal('dept-modal'); renderDepts();
  toast(id?'Department updated.':'Department added.'); loadDashboard();
}
async function deleteDept(id){
  if(!confirm('Delete this department?')) return;
  const r=await fetch(`/api/departments/${encodeURIComponent(id)}`,{method:'DELETE'});
  if(!r.ok){const d=await r.json();toast(d.error,'err');return;}
  await loadDepts(); renderDepts(); toast('Department deleted.'); loadDashboard();
}

// ── PROMOTIONS ─────────────────────────────────────────────────────────────
function renderPromos(){
  const tbody=document.getElementById('promo-tbody');
  if(!allPromos.length){
    tbody.innerHTML='<tr class="empty-row"><td colspan="5">No promotion requests yet.</td></tr>';
    return;
  }
  tbody.innerHTML=allPromos.map((p,i)=>`
    <tr data-key="${i}" onclick="selectPromo(${i},this)">
      <td><span style="color:var(--gold)">${esc(p.employee_id)}</span> ${esc(p.full_name||'')}</td>
      <td>${esc(p.current_level||'—')}</td>
      <td>${esc(p.requested_level||'—')}</td>
      <td>${esc(p.date||'—')}</td>
      <td><span class="badge badge-${p.status}">${p.status}</span></td>
    </tr>`).join('');
}
function selectPromo(idx,row){
  document.querySelectorAll('#promo-tbody tr').forEach(r=>r.classList.remove('selected'));
  row.classList.add('selected'); selPromoKey=idx;
}
async function submitPromo(){
  const empInput=document.getElementById('p-emp').value.trim();
  const lvl=document.getElementById('p-role').value.trim();
  if(!empInput||!lvl){toast('Employee and requested level are required.','err');return;}
  const r=await fetch('/api/promotions',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({employee_ref:empInput,requested_level:lvl})});
  const data=await r.json();
  if(!r.ok){toast(data.error,'err');return;}
  document.getElementById('p-emp').value='';
  document.getElementById('p-role').value='';
  await loadPromos(); await loadAudit(); renderPromos();
  toast('Promotion request submitted.'); loadDashboard();
}
async function resolvePromo(resolution){
  if(selPromoKey===null){toast('Select a request first.','err');return;}
  const p=allPromos[selPromoKey];
  if(p.status!=='pending'){toast('Request already resolved.','err');return;}
  const r=await fetch(`/api/promotions/${encodeURIComponent(p.employee_id)}/${p.req_index}`,
    {method:'PUT',headers:{'Content-Type':'application/json'},body:JSON.stringify({resolution})});
  if(!r.ok){const d=await r.json();toast(d.error,'err');return;}
  await loadPromos(); await loadEmployees(); await loadAudit();
  renderPromos(); selPromoKey=null;
  toast(`Request ${resolution==='approved'?'approved ✓':'denied ✕'}.`, resolution==='approved'?'ok':'err');
  loadDashboard();
}

// ── AUDIT ──────────────────────────────────────────────────────────────────
function renderAudit(){
  const q=(document.getElementById('audit-search')?.value||'').toLowerCase();
  const tbody=document.getElementById('audit-tbody');
  const list=auditLog.filter(a=>!q||
    [a.action,a.performed_by,a.target,a.details].some(v=>(v||'').toLowerCase().includes(q)));
  if(!list.length){tbody.innerHTML='<tr class="empty-row"><td colspan="5">No audit entries.</td></tr>';return;}
  tbody.innerHTML=list.map(a=>`
    <tr>
      <td style="color:var(--muted2)">${esc(a.timestamp)}</td>
      <td><span class="audit-action">${esc(a.action)}</span></td>
      <td><strong>${esc(a.performed_by)}</strong></td>
      <td>${esc(a.target||'—')}</td>
      <td style="color:var(--muted2)">${esc(a.details||'—')}</td>
    </tr>`).join('');
}
function filterAudit(){ renderAudit(); }

// ── SYSTEM USERS ───────────────────────────────────────────────────────────
function renderUsers(){
  const tbody=document.getElementById('users-tbody');
  const list=Object.entries(systemUsers);
  if(!list.length){tbody.innerHTML='<tr class="empty-row"><td colspan="6">No users.</td></tr>';return;}
  tbody.innerHTML=list.map(([u,d])=>`
    <tr>
      <td><strong>${esc(u)}</strong></td>
      <td>${esc(d.full_name||'—')}</td>
      <td><span class="badge badge-${d.role.toLowerCase()}">${esc(d.role)}</span></td>
      <td>${esc(d.employee_id||'—')}</td>
      <td>${esc(d.created_at||'—')}</td>
      <td>
        ${u!=='admin'?`<button class="btn btn-danger btn-sm" onclick="deleteUser('${esc(u)}')">✕ Delete</button>`:'<span style="color:var(--muted);font-size:11px">Protected</span>'}
      </td>
    </tr>`).join('');
}
function openAddUserModal(){
  ['um-uname','um-name','um-pw'].forEach(id=>document.getElementById(id).value='');
  document.getElementById('um-role').value='Staff';
  openModal('user-modal');
}
async function saveUser(){
  const uname=document.getElementById('um-uname').value.trim();
  const pw=document.getElementById('um-pw').value;
  if(!uname||!pw){toast('Username and password are required.','err');return;}
  const r=await fetch('/api/auth-users',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({username:uname,password:pw,
      full_name:document.getElementById('um-name').value.trim(),
      role:document.getElementById('um-role').value})});
  const data=await r.json();
  if(!r.ok){toast(data.error,'err');return;}
  await loadUsers(); closeModal('user-modal'); renderUsers();
  toast(`User '${uname}' created.`);
}
async function deleteUser(uname){
  if(!confirm(`Delete system user '${uname}'?`)) return;
  const r=await fetch(`/api/auth-users/${encodeURIComponent(uname)}`,{method:'DELETE'});
  if(!r.ok){const d=await r.json();toast(d.error,'err');return;}
  await loadUsers(); renderUsers(); toast('User deleted.');
}

// ── EXPORT ─────────────────────────────────────────────────────────────────
function exportData(fmt){ window.location.href=`/api/export/${fmt}`; }

// ── HELPERS ────────────────────────────────────────────────────────────────
function openModal(id){ document.getElementById(id).classList.add('open'); }
function closeModal(id){ document.getElementById(id).classList.remove('open'); }
document.querySelectorAll('.modal-overlay').forEach(o=>{
  o.addEventListener('click',e=>{ if(e.target===o) o.classList.remove('open'); });
});

function statusBadge(s){
  const map={'Active':'active','On Leave':'leave','Inactive':'inactive','Terminated':'terminated'};
  return `<span class="badge badge-${map[s]||'inactive'}">${esc(s||'Active')}</span>`;
}

function avatarHtml(e,size){
  if(e.avatar){
    return `<div class="av-mini" style="width:${size}px;height:${size}px"><img src="/avatars/${esc(e.avatar)}" alt="av"></div>`;
  }
  const init=(e.full_name||'?').split(' ').map(x=>x[0]).join('').slice(0,2).toUpperCase();
  return `<div class="av-mini" style="width:${size}px;height:${size}px;font-size:${Math.floor(size*.4)}px">${init}</div>`;
}

let _toastTimer=null;
function toast(msg,type='ok'){
  const el=document.getElementById('toast');
  el.textContent=msg; el.className='show '+(type==='err'?'err':'ok');
  if(_toastTimer) clearTimeout(_toastTimer);
  _toastTimer=setTimeout(()=>el.className='',3200);
}
function esc(s){ return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }
</script>
{% endraw %}
</body></html>"""

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — AUTH
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/")
def index():
    return redirect(url_for("login_page") if not get_current_user() else url_for("main_app"))

@app.route("/login", methods=["GET", "POST"])
def login_page():
    error = ""
    if request.method == "POST":
        uname = request.form.get("username", "").strip()
        pw    = request.form.get("password", "")
        d     = db.load_data()
        u     = d["auth_users"].get(uname)
        if u and db.verify_pw(pw, u["password_hash"], u["salt"]):
            session.permanent = True
            session["username"] = uname
            db.audit(d, "LOGIN", uname, details="Successful login")
            db.save_data(d)
            return redirect(url_for("main_app"))
        error = "Invalid username or password."
    return render_template_string(LOGIN_TPL, error=error, year=datetime.now().year)

@app.route("/logout")
def logout():
    u = get_current_user()
    if u:
        d = db.load_data()
        db.audit(d, "LOGOUT", u["username"])
        db.save_data(d)
    session.clear()
    return redirect(url_for("login_page"))

@app.route("/app")
@login_required
def main_app():
    u = get_current_user()
    return render_template_string(APP_TPL, user_json=json.dumps(u))

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — EMPLOYEES
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/api/employees")
@api_login_required
def get_employees():
    return jsonify(db.load_data()["employees"])

@app.route("/api/employees", methods=["POST"])
@perm_required("add_emp")
def add_employee():
    u = get_current_user()
    d = db.load_data()
    body = request.get_json() or {}
    if not body.get("full_name"):
        return jsonify({"error": "full_name is required"}), 400
    emp_id = db.next_emp_id(d)
    d["employees"][emp_id] = {
        "employee_id":  emp_id,
        "full_name":    body.get("full_name", ""),
        "email":        body.get("email", ""),
        "phone":        body.get("phone", ""),
        "age":          body.get("age", ""),
        "department":   body.get("department", ""),
        "position":     body.get("position", ""),
        "level":        body.get("level", "Junior"),
        "status":       body.get("status", "Active"),
        "location":     body.get("location", ""),
        "avatar":       None,
        "date_joined":  datetime.now().strftime("%Y-%m-%d"),
        "promotion_requests": [],
    }
    db.audit(d, "ADD_EMPLOYEE", u["username"], target=emp_id,
             details=f"Added {body.get('full_name')}")
    db.save_data(d)
    return jsonify({"employee_id": emp_id}), 201

@app.route("/api/employees/<emp_id>", methods=["PUT"])
@perm_required("edit_emp")
def update_employee(emp_id):
    u = get_current_user()
    d = db.load_data()
    if emp_id not in d["employees"]:
        return jsonify({"error": "Employee not found"}), 404
    body = request.get_json() or {}
    for key in ["full_name","email","phone","age","department","position","level","status","location"]:
        if key in body:
            d["employees"][emp_id][key] = body[key]
    db.audit(d, "EDIT_EMPLOYEE", u["username"], target=emp_id,
             details=f"Updated {d['employees'][emp_id].get('full_name')}")
    db.save_data(d)
    return jsonify({"ok": True})

@app.route("/api/employees/<emp_id>", methods=["DELETE"])
@perm_required("delete_emp")
def delete_employee(emp_id):
    u = get_current_user()
    d = db.load_data()
    if emp_id not in d["employees"]:
        return jsonify({"error": "Employee not found"}), 404
    name = d["employees"][emp_id].get("full_name", emp_id)
    del d["employees"][emp_id]
    db.audit(d, "DELETE_EMPLOYEE", u["username"], target=emp_id, details=f"Deleted {name}")
    db.save_data(d)
    return jsonify({"ok": True})

@app.route("/api/employees/<emp_id>/avatar", methods=["POST"])
@perm_required("edit_emp")
def upload_avatar(emp_id):
    d = db.load_data()
    if emp_id not in d["employees"]:
        return jsonify({"error": "Employee not found"}), 404
    if "avatar" not in request.files:
        return jsonify({"error": "No file"}), 400
    f   = request.files["avatar"]
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_IMG:
        return jsonify({"error": "Unsupported image type"}), 400
    fn = db.save_avatar(emp_id, f.read(), ext)
    d["employees"][emp_id]["avatar"] = fn
    db.save_data(d)
    return jsonify({"filename": fn})

@app.route("/avatars/<path:filename>")
def serve_avatar(filename):
    return send_from_directory(db.AVATAR_DIR, filename)

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — DEPARTMENTS
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/api/departments")
@api_login_required
def get_departments():
    return jsonify(db.load_data()["departments"])

@app.route("/api/departments", methods=["POST"])
@perm_required("manage_depts")
def add_department():
    u = get_current_user()
    d = db.load_data()
    body = request.get_json() or {}
    if not body.get("name"):
        return jsonify({"error": "name is required"}), 400
    existing = [v["name"].lower() for v in d["departments"].values()]
    if body["name"].lower() in existing:
        return jsonify({"error": "Department already exists"}), 409
    n = len(d["departments"]) + 1
    dept_id = f"DEPT-{n:03d}"
    while dept_id in d["departments"]:
        n += 1; dept_id = f"DEPT-{n:03d}"
    d["departments"][dept_id] = {
        "name": body["name"], "head": body.get("head", ""),
        "description": body.get("description", ""),
        "created_at": datetime.now().strftime("%Y-%m-%d"),
    }
    db.audit(d, "ADD_DEPARTMENT", u["username"], target=body["name"])
    db.save_data(d)
    return jsonify({"dept_id": dept_id}), 201

@app.route("/api/departments/<dept_id>", methods=["PUT"])
@perm_required("manage_depts")
def update_department(dept_id):
    u = get_current_user()
    d = db.load_data()
    if dept_id not in d["departments"]:
        return jsonify({"error": "Department not found"}), 404
    body = request.get_json() or {}
    for key in ["name", "head", "description"]:
        if key in body:
            d["departments"][dept_id][key] = body[key]
    db.audit(d, "EDIT_DEPARTMENT", u["username"], target=dept_id)
    db.save_data(d)
    return jsonify({"ok": True})

@app.route("/api/departments/<dept_id>", methods=["DELETE"])
@perm_required("manage_depts")
def delete_department(dept_id):
    u = get_current_user()
    d = db.load_data()
    if dept_id not in d["departments"]:
        return jsonify({"error": "Department not found"}), 404
    name = d["departments"][dept_id]["name"]
    del d["departments"][dept_id]
    db.audit(d, "DELETE_DEPARTMENT", u["username"], target=name)
    db.save_data(d)
    return jsonify({"ok": True})

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — PROMOTIONS
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/api/promotions")
@api_login_required
def get_promotions():
    d = db.load_data()
    rows = []
    for emp_id, e in d["employees"].items():
        for i, r in enumerate(e.get("promotion_requests", [])):
            rows.append({
                "employee_id":     emp_id,
                "full_name":       e.get("full_name", ""),
                "current_level":   r.get("current_level", ""),
                "requested_level": r.get("requested_level", ""),
                "date":            r.get("date", ""),
                "status":          r.get("status", "pending"),
                "req_index":       i,
            })
    return jsonify(rows)

@app.route("/api/promotions", methods=["POST"])
@api_login_required
def submit_promotion():
    u  = get_current_user()
    d  = db.load_data()
    b  = request.get_json() or {}
    ref = b.get("employee_ref", "").strip()
    lvl = b.get("requested_level", "").strip()
    if not ref or not lvl:
        return jsonify({"error": "employee_ref and requested_level required"}), 400
    # Find by ID or name
    emp_id = None
    if ref in d["employees"]:
        emp_id = ref
    else:
        for eid, e in d["employees"].items():
            if ref.lower() in e.get("full_name", "").lower():
                emp_id = eid; break
    if not emp_id:
        return jsonify({"error": f"Employee '{ref}' not found"}), 404
    req = {
        "requested_level": lvl,
        "current_level":   d["employees"][emp_id].get("level", ""),
        "status":          "pending",
        "date":            datetime.now().strftime("%Y-%m-%d"),
        "resolved_date":   "",
    }
    d["employees"][emp_id].setdefault("promotion_requests", []).append(req)
    db.audit(d, "PROMOTION_REQUEST", u["username"], target=emp_id,
             details=f"Requested: {lvl}")
    db.save_data(d)
    return jsonify({"ok": True}), 201

@app.route("/api/promotions/<emp_id>/<int:idx>", methods=["PUT"])
@perm_required("approve_promo")
def resolve_promotion(emp_id, idx):
    u = get_current_user()
    d = db.load_data()
    if emp_id not in d["employees"]:
        return jsonify({"error": "Employee not found"}), 404
    reqs = d["employees"][emp_id].get("promotion_requests", [])
    if idx >= len(reqs):
        return jsonify({"error": "Request not found"}), 404
    req = reqs[idx]
    if req["status"] != "pending":
        return jsonify({"error": "Already resolved"}), 409
    resolution = (request.get_json() or {}).get("resolution", "")
    if resolution not in ("approved", "denied"):
        return jsonify({"error": "resolution must be 'approved' or 'denied'"}), 400
    req["status"]        = resolution
    req["resolved_date"] = datetime.now().strftime("%Y-%m-%d")
    req["resolved_by"]   = u["username"]
    if resolution == "approved":
        d["employees"][emp_id]["level"] = req["requested_level"]
    db.audit(d, f"PROMOTION_{resolution.upper()}", u["username"],
             target=emp_id, details=f"→ {req['requested_level']}")
    db.save_data(d)
    return jsonify({"ok": True})

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — AUDIT / STATS
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/api/audit")
@perm_required("view_audit")
def get_audit():
    return jsonify(db.load_data()["audit_log"])

@app.route("/api/stats")
@api_login_required
def get_stats():
    return jsonify(db.compute_stats(db.load_data()))

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — SYSTEM USERS
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/api/auth-users")
@perm_required("manage_users")
def get_auth_users():
    d = db.load_data()
    safe = {u: {k: v for k, v in info.items()
                if k not in ("password_hash", "salt")}
            for u, info in d["auth_users"].items()}
    return jsonify(safe)

@app.route("/api/auth-users", methods=["POST"])
@perm_required("manage_users")
def add_auth_user():
    actor = get_current_user()
    d     = db.load_data()
    b     = request.get_json() or {}
    uname = b.get("username", "").strip()
    pw    = b.get("password", "")
    if not uname or not pw:
        return jsonify({"error": "username and password required"}), 400
    if uname in d["auth_users"]:
        return jsonify({"error": "Username already exists"}), 409
    h, s = db.hash_pw(pw)
    d["auth_users"][uname] = {
        "password_hash": h, "salt": s,
        "role":          b.get("role", "Staff"),
        "full_name":     b.get("full_name", ""),
        "employee_id":   b.get("employee_id"),
        "created_at":    datetime.now().strftime("%Y-%m-%d"),
    }
    db.audit(d, "ADD_USER", actor["username"], target=uname,
             details=f"Role: {b.get('role','Staff')}")
    db.save_data(d)
    return jsonify({"ok": True}), 201

@app.route("/api/auth-users/<username>", methods=["DELETE"])
@perm_required("manage_users")
def delete_auth_user(username):
    actor = get_current_user()
    d     = db.load_data()
    if username == "admin":
        return jsonify({"error": "Cannot delete the protected admin account"}), 403
    if username not in d["auth_users"]:
        return jsonify({"error": "User not found"}), 404
    del d["auth_users"][username]
    db.audit(d, "DELETE_USER", actor["username"], target=username)
    db.save_data(d)
    return jsonify({"ok": True})

# ═══════════════════════════════════════════════════════════════════════════════
#  ROUTES — EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
@app.route("/api/export/csv")
@login_required
def export_csv():
    d = db.load_data()
    headers = ["employee_id","full_name","email","phone","age",
               "department","position","level","status","location","date_joined"]
    buf = io.StringIO()
    w   = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    w.writeheader()
    for emp_id, e in d["employees"].items():
        row = {"employee_id": emp_id}
        row.update({k: e.get(k, "") for k in headers[1:]})
        w.writerow(row)
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),
                     mimetype="text/csv", as_attachment=True,
                     download_name="archer_employees.csv")

@app.route("/api/export/excel")
@login_required
def export_excel():
    if not EXCEL_OK:
        return "openpyxl not installed — run: pip install openpyxl", 500
    d  = db.load_data()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    hfill = PatternFill("solid", fgColor="04060d")
    hfont = Font(bold=True, color="C9A848", size=11)
    cols  = ["Employee ID","Full Name","Email","Phone","Age",
             "Department","Position","Level","Status","Location","Date Joined"]
    keys  = ["","full_name","email","phone","age",
             "department","position","level","status","location","date_joined"]
    for ci, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font=hfont; cell.fill=hfill
        cell.alignment=Alignment(horizontal="center")
    for ri, (eid, e) in enumerate(d["employees"].items(), 2):
        ws.cell(row=ri, column=1, value=eid)
        for ci, k in enumerate(keys[1:], 2):
            ws.cell(row=ri, column=ci, value=e.get(k,""))
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=18
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="archer_employees.xlsx")

# ═══════════════════════════════════════════════════════════════════════════════
#  RUN
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("\n  ╔══════════════════════════════════════╗")
    print("  ║   ARCHER ENTERPRISE — Web Edition   ║")
    print("  ╠══════════════════════════════════════╣")
    print("  ║   http://localhost:5000              ║")
    print("  ║   Default: admin / admin123          ║")
    print("  ╚══════════════════════════════════════╝\n")
    app.run(debug=True, port=5000)
